import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

os.makedirs("testing-reports", exist_ok=True)

MODULES = [
    "Authentication", "Dashboard", "Workout", "Diet", "Budget Diet",
    "Stretching", "Water Tracker", "Sleep Tracker", "Step Tracker", "AI Trainer",
    "Shop", "Cart", "Checkout", "Address", "Orders", "Profile",
    "Notifications", "Settings", "Localization", "Supabase Integration",
    "Responsive Web", "Android UI"
]

COLUMNS = [
    "Test Case ID", "Module", "Test Scenario", "Test Title", "Objective",
    "Preconditions", "Test Steps", "Test Data", "Expected Result",
    "Actual Result", "Priority", "Severity", "Status", "Remarks"
]

# ----------------------------------------------------
# Rich Category Scenario Templates
# ----------------------------------------------------
TEMPLATES = {
    "Selenium Web": [
        ("Verify responsive viewport layout grid scaling", "Launch Web Chrome/Edge driver at viewport resolution, verify container grid, element alignments and navbar collapse state.", "Viewport: {data_detail}", "UI layout scales cleanly without horizontal scrollbars or clipping."),
        ("Verify form input validation & error highlight rendering", "Focus input element, enter invalid boundary value, trigger blur event, inspect validation error badge.", "Input value: {data_detail}", "Validation hint appears immediately with accessible aria warning."),
        ("Verify interactive hover transitions & button clicks", "Hover cursor over CTA element, verify CSS transition state, trigger click action.", "Element selector: {data_detail}", "Button triggers transition and dispatches action without unhandled JS exceptions."),
        ("Verify modal backdrop blur, focus trap, and escape key close", "Trigger modal pop-up dialog, attempt keyboard tab traversal inside modal, hit ESC key.", "Modal dialog: {data_detail}", "Modal opens centered with focus trap; closes smoothly on ESC key event."),
        ("Verify data table sorting, search filtering, and pagination", "Navigate to list grid, enter search keyword, click column header to sort, navigate to next page.", "Page size = 15, Filter: {data_detail}", "Grid filters dynamically, sorts columns, and paginates smoothly."),
        ("Verify multi-tab window navigation & session persistence", "Open secondary browser tab, navigate to application route, perform action, return to tab 1.", "Session storage token: {data_detail}", "Authentication session persists seamlessly across browser tabs."),
        ("Verify web browser Back/Forward history state synchronization", "Deep link to feature screen, click Back button, then Forward button in browser.", "History stack: {data_detail}", "Browser history syncs cleanly with router path and DOM view."),
        ("Verify image lazy loading & WebP asset rendering", "Scroll down page rapidly to trigger lazy viewport loading of media content.", "Asset path: {data_detail}", "Images display placeholder skeleton before loading optimized WebP image asset."),
        ("Verify light/dark theme switching and CSS custom property updates", "Click global theme toggle in top header navbar.", "Theme toggle selector: {data_detail}", "CSS variables update globally across all visible DOM containers."),
        ("Verify keyboard tab order traversal and ARIA accessibility roles", "Press TAB key repeatedly from top of web page to traverse focusable DOM elements.", "Keyboard navigation: {data_detail}", "Focus ring highlights interactive elements in logical DOM order.")
    ],
    "Appium Android": [
        ("Verify native touch swipe & vertical list scrolling", "Execute vertical swipe gesture on scrollable RecyclerView/ListView container.", "W3C TouchAction: {data_detail}", "Container scrolls smoothly at 60fps without lag or touch drag locking."),
        ("Verify Android hardware Back button stack navigation", "Navigate deep into module screen hierarchy, press physical/software back button.", "KeyCode 4 (BACK): {data_detail}", "App pops top route cleanly and returns to previous view without exiting."),
        ("Verify screen auto-rotation (Portrait to Landscape)", "Rotate device orientation to landscape 90deg, inspect UI adaptation.", "Orientation: {data_detail}", "Layout resizes responsively in landscape mode without button overlap."),
        ("Verify native Android push notification banner & tap routing", "Trigger local push notification alert, open shade, tap notification banner.", "Payload ID: {data_detail}", "App opens target screen specified in push notification payload."),
        ("Verify Android BiometricPrompt fingerprint/face unlock", "Trigger security action requiring biometric confirmation.", "Biometric mock: {data_detail}", "Native Android BiometricPrompt appears; success unlocks target view."),
        ("Verify camera permission request dialog and avatar capture", "Tap upload photo action, handle Android system permission dialog, take photo.", "CAMERA permission: {data_detail}", "Permission granted dialog works; captured image updates target profile asset."),
        ("Verify Android deep link URI scheme execution via ADB", "Send ADB shell command to launch deep link URI into app.", "ADB intent: {data_detail}", "App opens directly into target screen specified in URI intent."),
        ("Verify app backgrounding, pause, and resume lifecycle state", "Press Home button to send app to background for 30s, re-open app.", "Android Lifecycle: {data_detail}", "App resumes state intact without memory leak or session timeout."),
        ("Verify system dark mode integration with Android OS settings", "Toggle system-wide Android Dark Theme switch from quick settings shade.", "OS Theme setting: {data_detail}", "NutriFit automatically adapts to system dark theme preference."),
        ("Verify offline cached state when switching Airplane mode", "Turn on Airplane Mode on device, navigate app screens.", "Offline network: {data_detail}", "App displays offline warning banner while serving cached local storage data.")
    ],
    "E2E": [
        ("Verify complete end-to-end user workflow execution", "Initiate multi-step workflow from entry point, complete all input forms, submit final transaction.", "Workflow payload: {data_detail}", "Entire end-to-end flow completes successfully and updates DB state."),
        ("Verify cross-module state synchronization after data mutation", "Mutate user data in one module, navigate to secondary module screen, inspect state.", "Sync payload: {data_detail}", "Secondary module reflects updated state immediately without manual refresh."),
        ("Verify multi-step wizard step navigation & back-tracking", "Progress to step 3 of wizard form, click back button to step 2, update field, proceed.", "Wizard state: {data_detail}", "Form state preserves entered values during forward and backward navigation."),
        ("Verify transactional data persistence & client sync", "Submit new record, trigger server refresh, inspect local database cache sync.", "DB Transaction: {data_detail}", "Record persists locally and syncs with Supabase PostgreSQL instance."),
        ("Verify error recovery and transaction rollbacks on failure", "Simulate server error mid-transaction, verify UI error boundary and state rollback.", "Fault injection: {data_detail}", "Transaction rolls back safely; clear user-friendly error toast displayed.")
    ],
    "Functional": [
        ("Verify component logic and state mutation functions", "Invoke component function with test parameters, observe return values and side effects.", "Function params: {data_detail}", "Function executes logic accurately and updates internal state correctly."),
        ("Verify validation constraints and edge-case handling", "Pass boundary, empty, and out-of-bounds inputs to validation handler.", "Edge case input: {data_detail}", "Validation rules enforce constraints and return precise error strings."),
        ("Verify asynchronous service calls and Futures resolution", "Execute async service method, await Future completion, parse response model.", "Async payload: {data_detail}", "Service parses JSON response cleanly into domain model instance."),
        ("Verify reactive state listener callbacks and UI rebuilds", "Mutate state property in Provider model, count listener callbacks triggered.", "Provider notify: {data_detail}", "Listeners receive notification and rebuild bound UI components once."),
        ("Verify authorization guard rules and permission checks", "Attempt feature access with restricted user permissions.", "Auth role: {data_detail}", "Authorization guard blocks access and redirects to permission error screen.")
    ],
    "Load": [
        ("Verify API endpoint response latency under 100 concurrent VUs", "Run Locust load generator targeting local server endpoint for 5 minutes.", "100 VU, Hatch rate 10/s: {data_detail}", "P95 latency < 300ms, HTTP error rate = 0.0%."),
        ("Verify UI rendering framerate (60 FPS) under continuous interaction", "Perform automated continuous UI actions while monitoring Flutter Performance overlay.", "FPS telemetry: {data_detail}", "Frame rendering stays steady at 60 FPS without jank or dropped frames."),
        ("Verify heap memory stability during 15-minute continuous test run", "Execute repeated view transitions while connected to Memory Profiler.", "Heap Profiler: {data_detail}", "Heap memory remains stable under 180MB without progressive memory leaks."),
        ("Verify bulk database batch insertion performance", "Insert 500 records concurrently into local Supabase Docker test instance.", "500 records: {data_detail}", "Total batch insertion execution completes in under 1.5 seconds."),
        ("Verify asset image loading optimization & cache hit speed", "Load media feed containing 50 WebP assets under throttled network conditions.", "Asset suite: {data_detail}", "First contentful paint < 400ms, cached image load < 50ms.")
    ],
    "Safe Code and Configuration Checks": [
        ("Verify environment variable isolation & absence of hardcoded secrets", "Audit repository codebase and build scripts for exposed plain-text credentials.", "Static Code Audit: {data_detail}", "Zero sensitive keys or credentials found hardcoded in source code."),
        ("Verify Supabase Row Level Security (RLS) enforcement", "Attempt SELECT/UPDATE on private user data using secondary user JWT token.", "RLS Policy Check: {data_detail}", "Database rejects unauthorized query with HTTP 403 / permission error."),
        ("Verify HTTPS/TLS 1.3 enforcement on all remote connection endpoints", "Inspect outbound network traffic using passive proxy audit tool.", "Passive OWASP Proxy: {data_detail}", "All HTTP connections upgrade to TLS 1.3/1.2; insecure calls blocked."),
        ("Verify SQL Injection sanitization on user input fields", "Submit payload `' OR '1'='1` in search and text input fields.", "Safe SQLi Payload: {data_detail}", "Input sanitized cleanly; no SQL syntax error or data leakage occurs."),
        ("Verify XSS (Cross-Site Scripting) escaping in rendered DOM content", "Submit payload `<script>alert('xss')</script>` into user text fields.", "Safe XSS Payload: {data_detail}", "Input escaped as literal text string; script tag not executed in DOM.")
    ]
}

# ----------------------------------------------------
# Generate 350 Test Cases for Category
# ----------------------------------------------------
def generate_350_category_cases(cat_name, prefix):
    cases = []
    scenarios = TEMPLATES[cat_name]

    for i in range(1, 351):
        tc_id = f"{prefix}-{i:03d}"
        mod = MODULES[(i - 1) % len(MODULES)]
        scen_tpl = scenarios[(i - 1) % len(scenarios)]

        detail_variant = f"Variant #{i} ({mod})"
        
        title = f"{cat_name} - {mod} - {scen_tpl[0]} (#{i})"
        scenario = f"{scen_tpl[0]} for {mod}"
        obj = f"Ensure {mod} module satisfies quality criteria for {cat_name} category."
        pre = f"Test environment active, NutriFit app launched on local test setup for {mod}."
        steps = f"1. Open NutriFit app on target setup.\n2. Navigate to {mod} screen.\n3. {scen_tpl[1]}\n4. Validate assertions."
        data = f"Module: {mod}, Setup: Local Host, Detail: {scen_tpl[2].format(data_detail=detail_variant)}"
        exp = f"{scen_tpl[3]} for module {mod}."
        
        prio = "High" if i % 3 == 0 else ("Low" if i % 5 == 0 else "Medium")
        sev = "Critical" if i % 7 == 0 else ("Major" if i % 2 == 0 else "Moderate")

        cases.append({
            "Test Case ID": tc_id,
            "Module": mod,
            "Test Scenario": scenario,
            "Test Title": title,
            "Objective": obj,
            "Preconditions": pre,
            "Test Steps": steps,
            "Test Data": data,
            "Expected Result": exp,
            "Actual Result": "Pending Execution",
            "Priority": prio,
            "Severity": sev,
            "Status": "Not Executed",
            "Remarks": f"Generated baseline case for {cat_name}"
        })

    return cases


# ----------------------------------------------------
# Build Workbook for Category
# ----------------------------------------------------
def build_workbook_for_category(filename, cat_name, cases):
    wb = openpyxl.Workbook()
    wb.remove(wb.active) # Remove default sheet

    font_family = "Segoe UI"
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="10A866", end_color="10A866", fill_type="solid") # Emerald Green

    title_font = Font(name=font_family, size=15, bold=True, color="0A5C36")
    subtitle_font = Font(name=font_family, size=10, italic=True, color="555555")
    section_font = Font(name=font_family, size=12, bold=True, color="0A5C36")

    thin_border = Border(
        left=Side(style='thin', color='E0E0E0'),
        right=Side(style='thin', color='E0E0E0'),
        top=Side(style='thin', color='E0E0E0'),
        bottom=Side(style='thin', color='E0E0E0')
    )

    card_border = Border(
        left=Side(style='medium', color='10A866'),
        right=Side(style='medium', color='10A866'),
        top=Side(style='medium', color='10A866'),
        bottom=Side(style='medium', color='10A866')
    )

    status_styles = {
        "Passed": (PatternFill(start_color="E8F8F0", end_color="E8F8F0", fill_type="solid"), Font(name=font_family, color="10A866", bold=True)),
        "Failed": (PatternFill(start_color="FCE8E6", end_color="FCE8E6", fill_type="solid"), Font(name=font_family, color="D93025", bold=True)),
        "Skipped": (PatternFill(start_color="FEF7E0", end_color="FEF7E0", fill_type="solid"), Font(name=font_family, color="B06000", bold=True)),
        "Blocked": (PatternFill(start_color="FFEFE6", end_color="FFEFE6", fill_type="solid"), Font(name=font_family, color="D9530F", bold=True)),
        "Not Executed": (PatternFill(start_color="F1F3F4", end_color="F1F3F4", fill_type="solid"), Font(name=font_family, color="5F6368", bold=True))
    }

    # ----------------------------------------------------
    # Sheet 1: Summary Dashboard
    # ----------------------------------------------------
    ws_sum = wb.create_sheet(title="Summary Dashboard")
    ws_sum.views.sheetView[0].showGridLines = True

    ws_sum["A1"] = f"NutriFit QA - {cat_name} Test Suite Summary"
    ws_sum["A1"].font = title_font
    ws_sum["A2"] = f"Category: {cat_name} | Total Test Cases: 350 | Target: Local Test Environment"
    ws_sum["A2"].font = subtitle_font

    ws_sum.row_dimensions[1].height = 24
    ws_sum.row_dimensions[2].height = 18

    # KPI Summary Cards
    kpis = [
        ("Total Test Cases", "=COUNT('Detailed Test Cases'!A:A)"),
        ("Passed", '=COUNTIF(\'Detailed Test Cases\'!M:M, "Passed")'),
        ("Failed", '=COUNTIF(\'Detailed Test Cases\'!M:M, "Failed")'),
        ("Skipped", '=COUNTIF(\'Detailed Test Cases\'!M:M, "Skipped")'),
        ("Blocked", '=COUNTIF(\'Detailed Test Cases\'!M:M, "Blocked")'),
        ("Not Executed", '=COUNTIF(\'Detailed Test Cases\'!M:M, "Not Executed")'),
        ("Pass Rate %", '=IF(B4=0, "0%", TEXT(B5/B4, "0.0%"))')
    ]

    for col_idx, (hdr, form) in enumerate(kpis, start=1):
        c_hdr = ws_sum.cell(row=4, column=col_idx, value=hdr)
        c_hdr.font = Font(name=font_family, size=9, bold=True, color="555555")
        c_hdr.alignment = Alignment(horizontal='center', vertical='center')
        c_hdr.fill = PatternFill(start_color="F5F7FA", end_color="F5F7FA", fill_type="solid")

        c_val = ws_sum.cell(row=5, column=col_idx, value=form)
        c_val.font = Font(name=font_family, size=14, bold=True, color="0A5C36")
        c_val.alignment = Alignment(horizontal='center', vertical='center')
        c_val.border = card_border

    # Module Coverage Summary Table
    ws_sum["A7"] = "Module-Wise Test Coverage Breakdown (All 22 Modules)"
    ws_sum["A7"].font = section_font

    mod_headers = ["Module Name", "Total Cases", "Passed", "Failed", "Skipped", "Blocked", "Not Executed"]
    ws_sum.append(mod_headers)
    ws_sum.row_dimensions[8].height = 24

    for c_idx in range(1, len(mod_headers) + 1):
        cell = ws_sum.cell(row=8, column=c_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for r_offset, mod in enumerate(MODULES, start=9):
        ws_sum.cell(row=r_offset, column=1, value=mod)
        ws_sum.cell(row=r_offset, column=2, value=f'=COUNTIF(\'Detailed Test Cases\'!B:B, A{r_offset})')
        ws_sum.cell(row=r_offset, column=3, value=f'=COUNTIFS(\'Detailed Test Cases\'!B:B, A{r_offset}, \'Detailed Test Cases\'!M:M, "Passed")')
        ws_sum.cell(row=r_offset, column=4, value=f'=COUNTIFS(\'Detailed Test Cases\'!B:B, A{r_offset}, \'Detailed Test Cases\'!M:M, "Failed")')
        ws_sum.cell(row=r_offset, column=5, value=f'=COUNTIFS(\'Detailed Test Cases\'!B:B, A{r_offset}, \'Detailed Test Cases\'!M:M, "Skipped")')
        ws_sum.cell(row=r_offset, column=6, value=f'=COUNTIFS(\'Detailed Test Cases\'!B:B, A{r_offset}, \'Detailed Test Cases\'!M:M, "Blocked")')
        ws_sum.cell(row=r_offset, column=7, value=f'=COUNTIFS(\'Detailed Test Cases\'!B:B, A{r_offset}, \'Detailed Test Cases\'!M:M, "Not Executed")')

        for c_idx in range(1, 8):
            cell = ws_sum.cell(row=r_offset, column=c_idx)
            cell.font = Font(name=font_family, size=9)
            cell.border = thin_border
            if c_idx >= 2:
                cell.alignment = Alignment(horizontal='center', vertical='center')

    # Total Row
    tot_row = 9 + len(MODULES)
    ws_sum.cell(row=tot_row, column=1, value="Total Suite").font = Font(name=font_family, size=10, bold=True)
    for c_idx in range(2, 8):
        col_let = get_column_letter(c_idx)
        cell = ws_sum.cell(row=tot_row, column=c_idx, value=f"=SUM({col_let}9:{col_let}{tot_row-1})")
        cell.font = Font(name=font_family, size=10, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        cell.fill = PatternFill(start_color="F0F9F4", end_color="F0F9F4", fill_type="solid")

    ws_sum.column_dimensions['A'].width = 30
    for c in ['B', 'C', 'D', 'E', 'F', 'G']:
        ws_sum.column_dimensions[c].width = 16

    # ----------------------------------------------------
    # Sheet 2: Detailed Test Cases
    # ----------------------------------------------------
    ws_det = wb.create_sheet(title="Detailed Test Cases")
    ws_det.views.sheetView[0].showGridLines = True
    ws_det.append(COLUMNS)
    ws_det.row_dimensions[1].height = 28
    ws_det.freeze_panes = "A2"

    for c_idx in range(1, len(COLUMNS) + 1):
        cell = ws_det.cell(row=1, column=c_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for r_idx, tc in enumerate(cases, start=2):
        row_vals = [tc[c] for c in COLUMNS]
        ws_det.append(row_vals)
        ws_det.row_dimensions[r_idx].height = 36

        for c_idx in range(1, len(COLUMNS) + 1):
            cell = ws_det.cell(row=r_idx, column=c_idx)
            cell.font = Font(name=font_family, size=9)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='top', wrap_text=True)

            if c_idx in [1, 11, 12, 13]: # ID, Priority, Severity, Status
                cell.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)

            # Apply Status Style (13th column)
            if c_idx == 13:
                status_val = cell.value
                if status_val in status_styles:
                    fill, font = status_styles[status_val]
                    cell.fill = fill
                    cell.font = font

    ws_det.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{len(cases) + 1}"

    col_widths = {
        "Test Case ID": 14,
        "Module": 20,
        "Test Scenario": 28,
        "Test Title": 35,
        "Objective": 35,
        "Preconditions": 30,
        "Test Steps": 45,
        "Test Data": 28,
        "Expected Result": 35,
        "Actual Result": 20,
        "Priority": 12,
        "Severity": 12,
        "Status": 15,
        "Remarks": 25
    }

    for c_idx, c_name in enumerate(COLUMNS, start=1):
        ws_det.column_dimensions[get_column_letter(c_idx)].width = col_widths.get(c_name, 20)

    # ----------------------------------------------------
    # Sheet 3: Execution Evidence
    # ----------------------------------------------------
    ws_ev = wb.create_sheet(title="Execution Evidence")
    ws_ev.views.sheetView[0].showGridLines = True
    ev_cols = ["Test Case ID", "Module", "Test Scenario", "Status", "Evidence File Path", "Timestamp", "Verification Notes"]
    ws_ev.append(ev_cols)
    ws_ev.row_dimensions[1].height = 28
    ws_ev.freeze_panes = "A2"

    for c_idx in range(1, len(ev_cols) + 1):
        cell = ws_ev.cell(row=1, column=c_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for idx, tc in enumerate(cases[:15], start=2):
        ws_ev.append([
            tc["Test Case ID"],
            tc["Module"],
            tc["Test Scenario"],
            tc["Status"],
            f"testing-reports/evidence/{tc['Test Case ID'].lower()}_log.png",
            "2026-07-21 23:45:00 UTC",
            "Baseline artifact placeholder initialized"
        ])
        for c_idx in range(1, len(ev_cols) + 1):
            cell = ws_ev.cell(row=idx, column=c_idx)
            cell.font = Font(name=font_family, size=9)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='top', wrap_text=True)

    ws_ev.auto_filter.ref = f"A1:{get_column_letter(len(ev_cols))}16"
    for c_idx in range(1, len(ev_cols) + 1):
        ws_ev.column_dimensions[get_column_letter(c_idx)].width = 24

    filepath = os.path.join("testing-reports", filename)
    wb.save(filepath)
    print(f"SUCCESS: Generated '{filepath}' with 350 test cases.")


# ----------------------------------------------------
# Main Execution Entrypoint
# ----------------------------------------------------
def main():
    category_configs = [
        ("Selenium_Web_Test_Cases.xlsx", "Selenium Web", "SEL"),
        ("Appium_Android_Test_Cases.xlsx", "Appium Android", "APP"),
        ("E2E_Test_Cases.xlsx", "E2E", "E2E"),
        ("Functional_Test_Cases.xlsx", "Functional", "FUN"),
        ("Load_Test_Cases.xlsx", "Load", "LOD"),
        ("Safe_Code_and_Configuration_Checks.xlsx", "Safe Code and Configuration Checks", "SEC")
    ]

    total_all_workbooks = 0

    for filename, cat_name, prefix in category_configs:
        cases = generate_350_category_cases(cat_name, prefix)
        assert len(cases) == 350, f"Error generating 350 cases for {cat_name}"
        build_workbook_for_category(filename, cat_name, cases)
        total_all_workbooks += len(cases)

    print(f"\n==================================================")
    print(f"TOTAL TEST CASES GENERATED ACROSS ALL 6 FILES: {total_all_workbooks}")
    print(f"==================================================")

if __name__ == "__main__":
    main()
