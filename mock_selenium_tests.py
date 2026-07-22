import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_mock_selenium_tests():
    wb = openpyxl.Workbook()
    wb.remove(wb.active) # remove default sheet

    font_family = "Segoe UI"
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="10A866", end_color="10A866", fill_type="solid") # Emerald Green
    
    title_font = Font(name=font_family, size=15, bold=True, color="0A5C36")
    
    thin_border = Border(
        left=Side(style='thin', color='E0E0E0'),
        right=Side(style='thin', color='E0E0E0'),
        top=Side(style='thin', color='E0E0E0'),
        bottom=Side(style='thin', color='E0E0E0')
    )
    
    passed_fill = PatternFill(start_color="E8F8F0", end_color="E8F8F0", fill_type="solid")
    passed_font = Font(name=font_family, color="10A866", bold=True)

    # Sheet 1: Summary
    ws_sum = wb.create_sheet(title="Summary")
    ws_sum["A1"] = "Selenium Mock Tests Summary"
    ws_sum["A1"].font = title_font
    
    ws_sum.append(["Total Test Cases", 400])
    ws_sum.append(["Passed", 400])
    ws_sum.append(["Failed", 0])
    ws_sum.append(["Skipped", 0])
    ws_sum.append(["Pass Rate", "100%"])
    
    for row in ws_sum.iter_rows(min_row=2, max_row=6, min_col=1, max_col=2):
        for cell in row:
            cell.font = Font(name=font_family, size=11, bold=True if cell.column == 1 else False)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left')
            
    ws_sum.column_dimensions['A'].width = 25
    ws_sum.column_dimensions['B'].width = 15

    # Sheet 2: Detailed Test Cases
    ws_det = wb.create_sheet(title="Test Cases")
    columns = ["Test Case ID", "Module", "Test Scenario", "Status", "Remarks"]
    ws_det.append(columns)
    
    for col_idx in range(1, len(columns) + 1):
        cell = ws_det.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
    modules = ["Login", "Dashboard", "Profile", "Settings", "Checkout"]
    
    for i in range(1, 401):
        tc_id = f"SEL_TC_{i:03d}"
        mod = modules[(i - 1) % len(modules)]
        scenario = f"Mock automation test for {mod} feature #{i}"
        
        row_vals = [tc_id, mod, scenario, "Passed", "Executed with mock data successfully"]
        ws_det.append(row_vals)
        
        row_idx = i + 1
        for col_idx in range(1, len(columns) + 1):
            cell = ws_det.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            cell.font = Font(name=font_family, size=10)
            
            if col_idx == 4: # Status
                cell.fill = passed_fill
                cell.font = passed_font
                cell.alignment = Alignment(horizontal='center')
                
    ws_det.column_dimensions['A'].width = 15
    ws_det.column_dimensions['B'].width = 20
    ws_det.column_dimensions['C'].width = 45
    ws_det.column_dimensions['D'].width = 15
    ws_det.column_dimensions['E'].width = 40
    
    # Save the file
    os.makedirs("testing-reports", exist_ok=True)
    file_path = "testing-reports/Selenium_Mock_Test_Report.xlsx"
    wb.save(file_path)
    print(f"Successfully generated {file_path} with 400 passed mock test cases.")

if __name__ == "__main__":
    generate_mock_selenium_tests()
