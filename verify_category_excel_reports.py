import os
import openpyxl

category_files = [
    ("Selenium_Web_Test_Cases.xlsx", "Selenium Web"),
    ("Appium_Android_Test_Cases.xlsx", "Appium Android"),
    ("E2E_Test_Cases.xlsx", "E2E"),
    ("Functional_Test_Cases.xlsx", "Functional"),
    ("Load_Test_Cases.xlsx", "Load"),
    ("Safe_Code_and_Configuration_Checks.xlsx", "Safe Code and Configuration Checks")
]

REQUIRED_MODULES = [
    "Authentication", "Dashboard", "Workout", "Diet", "Budget Diet",
    "Stretching", "Water Tracker", "Sleep Tracker", "Step Tracker", "AI Trainer",
    "Shop", "Cart", "Checkout", "Address", "Orders", "Profile",
    "Notifications", "Settings", "Localization", "Supabase Integration",
    "Responsive Web", "Android UI"
]

EXPECTED_COLUMNS = [
    "Test Case ID", "Module", "Test Scenario", "Test Title", "Objective",
    "Preconditions", "Test Steps", "Test Data", "Expected Result",
    "Actual Result", "Priority", "Severity", "Status", "Remarks"
]

grand_total_cases = 0

print("=== STARTING RIGOROUS VERIFICATION OF ALL 6 CATEGORY EXCEL WORKBOOKS ===\n")

for filename, cat_name in category_files:
    filepath = os.path.join("testing-reports", filename)
    assert os.path.exists(filepath), f"File missing: {filepath}"

    wb = openpyxl.load_workbook(filepath, data_only=False)
    assert "Detailed Test Cases" in wb.sheetnames, f"Sheet 'Detailed Test Cases' missing in {filename}"
    assert "Summary Dashboard" in wb.sheetnames, f"Sheet 'Summary Dashboard' missing in {filename}"

    ws_det = wb["Detailed Test Cases"]
    rows = list(ws_det.iter_rows(values_only=True))

    header = rows[0]
    data_rows = rows[1:]

    # Validate Columns
    assert list(header) == EXPECTED_COLUMNS, f"Columns mismatch in {filename}.\nExpected: {EXPECTED_COLUMNS}\nGot: {list(header)}"

    tc_count = len(data_rows)
    print(f"File '{filename}': Found {tc_count} test cases.")
    assert tc_count == 350, f"Expected 350 test cases in {filename}, got {tc_count}"

    ids_in_file = set()
    modules_in_file = set()

    for r_idx, row in enumerate(data_rows, start=2):
        tc_id = row[0]
        mod = row[1]
        status = row[12]

        assert tc_id not in ids_in_file, f"Duplicate Test ID '{tc_id}' found at row {r_idx} in {filename}"
        ids_in_file.add(tc_id)
        modules_in_file.add(mod)

        assert status == "Not Executed", f"Invalid status '{status}' for {tc_id} in {filename}. Must be 'Not Executed'"

    assert len(ids_in_file) == 350, f"ID count mismatch in {filename}"
    grand_total_cases += tc_count

    missing_mods = [m for m in REQUIRED_MODULES if m not in modules_in_file]
    print(f"   -> Unique IDs: {len(ids_in_file)} | Modules Covered: {len(modules_in_file)}/22 | Status: All 'Not Executed'")
    assert len(missing_mods) == 0, f"Missing modules in {filename}: {missing_mods}"

print("\n==================================================")
print(f"VERIFICATION SUCCESSFUL: GRAND TOTAL = {grand_total_cases} TEST CASES!")
print("ALL 6 EXCEL WORKBOOKS VALIDATED PERFECTLY WITH ZERO ERRORS.")
print("==================================================")
