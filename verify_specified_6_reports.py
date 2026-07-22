import os
import openpyxl

SPECIFIED_FILES = [
    ("Selenium", "NutriFit_Selenium_Test_Report.xlsx"),
    ("Appium", "NutriFit_Appium_Test_Report.xlsx"),
    ("E2E", "NutriFit_E2E_Test_Report.xlsx"),
    ("Functional", "NutriFit_Functional_Test_Report.xlsx"),
    ("Load", "NutriFit_Load_Test_Report.xlsx"),
    ("Vulnerability", "NutriFit_Vulnerability_Test_Report.xlsx")
]

EXPECTED_MODULES = [
    "SplashScreen", "LoginScreen", "SignUpScreen", "ForgotScreen", "SelectScreen",
    "GoalScreen", "GenderScreen", "FoodScreen", "LocationScreen", "NumberScreen",
    "DashboardScreen", "HomeTab", "HeroCard", "WorkoutDayCard", "WarmupCard",
    "DietDayCard", "BudgetFoodSection", "PlansTab", "TrackersTab", "WaterTrackerSection",
    "SleepTrackerSection", "StepTrackerSection", "AITrainerScreen", "ShopTab", "BudgetProductCard",
    "CartScreen", "CheckoutScreen", "AddressManagementSection", "OrderConfirmationScreen", "MyOrdersScreen",
    "OrderTrackingTimeline", "ProfileTab", "RemindersNotificationModule", "SettingsLocalizationModule", "SupabaseIntegrationModule"
]

EXPECTED_COLUMNS = [
    "Test Case ID", "Module", "Test Scenario", "Test Steps",
    "Expected Result", "Actual Result", "Status", "Priority"
]

def verify_all_workbooks():
    num_modules = len(EXPECTED_MODULES)
    expected_cases_per_file = num_modules * 10
    expected_total_cases = expected_cases_per_file * len(SPECIFIED_FILES)

    total_duplicates_id = 0
    total_duplicates_scenario = 0
    total_missing = 0
    grand_total_cases = 0
    file_counts = {}

    print("=== STARTING PROGRAMMATIC VERIFICATION OF ALL 6 EXCEL WORKBOOKS ===\n")

    for cat_label, filename in SPECIFIED_FILES:
        rep_path = os.path.join("testing-reports", filename)
        assert os.path.exists(rep_path), f"File missing in testing-reports/: {rep_path}"
        assert os.path.exists(filename), f"File missing in root: {filename}"

        # 1. Verify workbook opens successfully
        wb = openpyxl.load_workbook(rep_path, data_only=False)
        assert "Summary Dashboard" in wb.sheetnames, f"Sheet 'Summary Dashboard' missing in {filename}"
        assert "Detailed Test Cases" in wb.sheetnames, f"Sheet 'Detailed Test Cases' missing in {filename}"
        assert "Execution Evidence" in wb.sheetnames, f"Sheet 'Execution Evidence' missing in {filename}"

        ws_det = wb["Detailed Test Cases"]
        rows = list(ws_det.iter_rows(values_only=True))

        header = rows[0]
        data_rows = rows[1:]

        # 2. Verify all required columns are present in exact order
        assert list(header) == EXPECTED_COLUMNS, f"Columns mismatch in {filename}.\nExpected: {EXPECTED_COLUMNS}\nGot: {list(header)}"

        tc_count = len(data_rows)
        file_counts[cat_label] = tc_count

        # 3. Verify total count matches actual_screen_count x 10
        if tc_count != expected_cases_per_file:
            print(f"ERROR: {filename} contains {tc_count} test cases (Expected: {expected_cases_per_file})")
            total_missing += abs(expected_cases_per_file - tc_count)

        expected_ids = {f"TC_{i:03d}" for i in range(1, expected_cases_per_file + 1)}
        ids_in_file = set()
        scenarios_in_file = set()
        module_counts = {mod: 0 for mod in EXPECTED_MODULES}
        dup_ids_in_file = 0
        dup_scenarios_in_file = 0

        for r_idx, row in enumerate(data_rows, start=2):
            tc_id = row[0]
            mod_name = row[1]
            scenario = row[2]
            actual_result = row[5]
            status = row[6]

            # 4. Check duplicate Test Case IDs
            if tc_id in ids_in_file:
                dup_ids_in_file += 1
            else:
                ids_in_file.add(tc_id)

            # 5. Check duplicate Test Scenarios
            if scenario in scenarios_in_file:
                dup_scenarios_in_file += 1
            else:
                scenarios_in_file.add(scenario)

            if mod_name in module_counts:
                module_counts[mod_name] += 1

            assert actual_result == "Not Executed", f"Invalid Actual Result '{actual_result}' for {tc_id} in {filename}."
            assert status == "Not Executed", f"Invalid Status '{status}' for {tc_id} in {filename}."

        # 6. Verify every actual screen/module has EXACTLY 10 test cases
        for mod, count in module_counts.items():
            assert count == 10, f"Module '{mod}' in {filename} has {count} test cases, expected 10."

        missing_ids = expected_ids - ids_in_file
        total_duplicates_id += dup_ids_in_file
        total_duplicates_scenario += dup_scenarios_in_file
        total_missing += len(missing_ids)

        assert dup_ids_in_file == 0, f"Found {dup_ids_in_file} duplicate IDs in {filename}"
        assert dup_scenarios_in_file == 0, f"Found {dup_scenarios_in_file} duplicate scenarios in {filename}"
        assert len(missing_ids) == 0, f"Missing IDs in {filename}: {missing_ids}"

        grand_total_cases += tc_count

    validation_passed = (
        grand_total_cases == expected_total_cases and
        total_duplicates_id == 0 and
        total_duplicates_scenario == 0 and
        total_missing == 0 and
        all(c == expected_cases_per_file for c in file_counts.values())
    )

    print("\n==================================================")
    print("NUTRIFIT SYSTEM SCREEN & MODULE TESTING VALIDATION")
    print("==================================================")
    print(f"Total screens/modules found: {num_modules}")
    print("10 test cases per screen/module: YES")
    print(f"Test cases per Excel file: {expected_cases_per_file}")
    print(f"Total test cases across all 6 files: {grand_total_cases}")
    print(f"Duplicate Test Case ID count: {total_duplicates_id}")
    print(f"Duplicate Test Scenario count: {total_duplicates_scenario}")
    print(f"Missing Test Case ID count: {total_missing}")
    print(f"Number of Excel files generated: {len(SPECIFIED_FILES)}")
    print("\nExcel Files Summary:")
    for idx, (cat_label, filename) in enumerate(SPECIFIED_FILES, start=1):
        print(f"{idx}. {filename}: {file_counts[cat_label]} test cases ({num_modules} screens x 10)")
    print(f"\nValidation: {'PASSED' if validation_passed else 'FAILED'} (10 test cases x {num_modules} screens x 6 testing categories = {expected_total_cases} total)")
    print("==================================================")

    if not validation_passed:
        raise ValueError("Validation failed! Not all workbooks meet the strict criteria.")

if __name__ == "__main__":
    verify_all_workbooks()
