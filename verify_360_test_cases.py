import openpyxl

file_path = "testing-reports/NutriFit_360_Test_Cases.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=False)

print("Sheet Names:", wb.sheetnames)

EXPECTED_SHEETS = [
    "Test Summary", "Selenium Tests", "Appium E2E Tests", "Flutter Tests",
    "Functional Tests", "Load Tests", "Safe Vulnerability Checks",
    "Failed Tests", "Execution Evidence"
]

for sheet in EXPECTED_SHEETS:
    assert sheet in wb.sheetnames, f"Missing sheet: {sheet}"

detail_sheet_counts = {
    "Selenium Tests": 100,
    "Appium E2E Tests": 90,
    "Flutter Tests": 60,
    "Functional Tests": 50,
    "Load Tests": 30,
    "Safe Vulnerability Checks": 30
}

all_ids = []
all_modules = set()
total_cases = 0

for sheet_name, expected_count in detail_sheet_counts.items():
    ws = wb[sheet_name]
    # Header is row 1
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    data_rows = rows[1:]

    actual_count = len(data_rows)
    print(f"Sheet '{sheet_name}': Expected {expected_count}, Got {actual_count}")
    assert actual_count == expected_count, f"Mismatch in {sheet_name}: expected {expected_count}, got {actual_count}"

    for row in data_rows:
        tc_id = row[0]
        tt_type = row[1]
        module = row[2]
        status = row[13]

        all_ids.append(tc_id)
        all_modules.add(module)
        assert status == "Not Executed", f"Invalid initial status '{status}' for test case {tc_id}"
        total_cases += 1

print(f"\nTotal Test Cases Verified Across Detail Sheets: {total_cases}")
assert total_cases == 360, f"Expected 360 total test cases, got {total_cases}"

# Check for duplicates
unique_ids = set(all_ids)
print(f"Total Unique Test Case IDs: {len(unique_ids)}")
assert len(unique_ids) == 360, f"Duplicate Test Case IDs found! {len(all_ids) - len(unique_ids)} duplicates."

# Check 37 modules
REQUIRED_MODULES = [
    "Splash", "Login", "Signup", "Forgot Password", "Email Verification",
    "Onboarding", "Dashboard", "Profile", "Language Settings", "AI Trainer",
    "Workout Plan", "Diet Plan", "Budget Diet Plan", "Stretching and Warm-up",
    "Water Tracker", "Sleep Tracker", "Step Tracker", "Calorie Calculator",
    "Meal Reminder", "Notifications", "Shop", "Product Search and Filters",
    "Product Details", "Wishlist", "Cart", "Checkout", "Address Management",
    "Payment Method", "Place Order", "My Orders", "Order Details",
    "Order Cancellation", "Supabase Integration", "Responsive Web UI",
    "Android UI", "Error Handling", "Logout"
]

missing_modules = [m for m in REQUIRED_MODULES if m not in all_modules]
print(f"Modules Covered: {len(all_modules)} / {len(REQUIRED_MODULES)}")
assert len(missing_modules) == 0, f"Missing modules: {missing_modules}"

print("\nALL VERIFICATION ASSERTS PASSED PERFECTLY!")
