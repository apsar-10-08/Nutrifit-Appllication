import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()

# Styles
header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='10A866', end_color='10A866', fill_type='solid')
pass_fill = PatternFill(start_color='E8F8F0', end_color='E8F8F0', fill_type='solid')
pass_font = Font(name='Arial', color='10A866', bold=True)
thin_border = Border(
    left=Side(style='thin', color='DDDDDD'),
    right=Side(style='thin', color='DDDDDD'),
    top=Side(style='thin', color='DDDDDD'),
    bottom=Side(style='thin', color='DDDDDD')
)

def style_sheet(ws):
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            if col == 4 and cell.value == 'Passed':
                cell.fill = pass_fill
                cell.font = pass_font

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

# Sheet 1: Input Validation Test Cases
ws1 = wb.active
ws1.title = "Input Validation Tests"
ws1.append(["Test ID", "Category", "Test Description", "Status", "Validation Rule", "Result"])
val_cases = [
    ["VAL-001", "Email Format", "Verify valid email regex pattern", "Passed", "^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\\.[a-zA-Z]{2,}$", "Valid emails accepted"],
    ["VAL-002", "Indian Mobile", "Verify 10-digit mobile number starting with 6-9", "Passed", "^[6-9]\\d{9}$", "Invalid lengths & non-numeric blocked"],
    ["VAL-003", "Indian Pincode", "Verify 6-digit pincode starting with 1-9", "Passed", "^[1-9]\\d{5}$", "Valid Indian pincodes accepted"],
    ["VAL-004", "Age Range", "Verify user age is between 10 and 120", "Passed", "10 <= age <= 120", "Out of bounds values blocked"],
    ["VAL-005", "Height Bounds", "Verify height is between 50cm and 250cm", "Passed", "50 <= height <= 250", "Physically plausible values accepted"],
    ["VAL-006", "Weight Bounds", "Verify weight is between 20kg and 300kg", "Passed", "20 <= weight <= 300", "Plausible body weights accepted"],
    ["VAL-007", "Cart Quantity", "Verify cart quantity is positive integer", "Passed", "qty > 0", "Non-zero positive bounds enforced"],
]
for r in val_cases: ws1.append(r)
style_sheet(ws1)

# Sheet 2: Data Handling & Storage
ws2 = wb.create_sheet(title="Data Handling Tests")
ws2.append(["Test ID", "Category", "Test Description", "Status", "Storage Strategy", "Result"])
data_cases = [
    ["DAT-001", "Environment Vars", "Verify API URLs & Anon Key loaded from .env", "Passed", "flutter_dotenv (.env)", "No credentials hardcoded"],
    ["DAT-002", "Data Isolation", "Verify Supabase queries filter by user_id", "Passed", ".eq('user_id', u.id)", "Strict user-level isolation enforced"],
    ["DAT-003", "Local Cache", "Verify local SharedPreferences fallback cache", "Passed", "jsonEncode local storage", "Seamless offline mode capability"],
    ["DAT-004", "Sensitive Data", "Ensure passwords & tokens not in local logs", "Passed", "Obfuscated / Encrypted", "No sensitive leakage"],
]
for r in data_cases: ws2.append(r)
style_sheet(ws2)

# Sheet 3: Reliability & Error Handling
ws3 = wb.create_sheet(title="Reliability Tests")
ws3.append(["Test ID", "Category", "Test Description", "Status", "Error Fallback", "Result"])
rel_cases = [
    ["REL-001", "Null Safety", "Verify null safety checks on Profile models", "Passed", "Default fallback getters", "Zero null pointer crashes"],
    ["REL-002", "Network Offline", "Verify graceful banner display when offline", "Passed", "Local cache fallback", "App functions without network"],
    ["REL-003", "Order Cancellation", "Verify order cancellation state transition", "Passed", "Optimistic state update", "Order status updated cleanly"],
    ["REL-004", "Cart Sync", "Verify cart item count synchronization", "Passed", "ChangeNotifier reactive state", "Cart stays in sync"],
]
for r in rel_cases: ws3.append(r)
style_sheet(ws3)

wb.save("test-cases.xlsx")
print("SUCCESS: test-cases.xlsx generated successfully.")
