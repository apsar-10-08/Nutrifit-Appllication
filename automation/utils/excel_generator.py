import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from automation.config import settings

def generate_excel_reports(test_results):
    passed_list = [t for t in test_results if t['status'] == 'Passed']
    failed_list = [t for t in test_results if t['status'] == 'Failed']
    skipped_list = [t for t in test_results if t['status'] in ['Skipped', 'Blocked', 'Pending']]

    total_tests = len(test_results)
    passed_count = len(passed_list)
    failed_count = len(failed_list)
    skipped_count = len(skipped_list)
    success_rate = (passed_count / total_tests * 100) if total_tests > 0 else 0
    total_duration = sum(t.get('duration', 0.0) for t in test_results)

    # Styles
    font_header = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    fill_header = PatternFill(start_color='10A866', end_color='10A866', fill_type='solid') # NutriFit Green
    font_bold = Font(name='Arial', size=11, bold=True)
    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    
    thin_border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD')
    )

    def style_row(ws, row_idx, font=None, fill=None, align=None, border=thin_border):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col)
            if font: cell.font = font
            if fill: cell.fill = fill
            if align: cell.alignment = align
            if border: cell.border = border

    def auto_fit_columns(ws):
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # 1. Main Automation_Test_Report.xlsx
    wb = Workbook()
    
    # Sheet 1: Executed Test Cases
    ws_executed = wb.active
    ws_executed.title = "Executed Test Cases"
    headers_executed = ["Test ID", "Module", "Test Name", "Status", "Execution Time (s)", "Priority"]
    ws_executed.append(headers_executed)
    style_row(ws_executed, 1, font=font_header, fill=fill_header, align=align_center)

    for t in test_results:
        ws_executed.append([t['id'], t['category'], t['name'], t['status'], round(t.get('duration', 0.0), 3), t['priority']])
        # Style status cells dynamically
        status_cell = ws_executed.cell(row=ws_executed.max_row, column=4)
        if t['status'] == 'Passed':
            status_cell.fill = PatternFill(start_color='E8F8F0', end_color='E8F8F0', fill_type='solid')
            status_cell.font = Font(color='10A866', bold=True)
        elif t['status'] == 'Failed':
            status_cell.fill = PatternFill(start_color='FFEEEE', end_color='FFEEEE', fill_type='solid')
            status_cell.font = Font(color='FF3333', bold=True)
        style_row(ws_executed, ws_executed.max_row, border=thin_border)
    auto_fit_columns(ws_executed)

    # Sheet 2: Passed Tests
    ws_passed = wb.create_sheet(title="Passed Tests")
    ws_passed.append(["Test ID", "Module", "Test Name", "Execution Time (s)", "Priority"])
    style_row(ws_passed, 1, font=font_header, fill=fill_header, align=align_center)
    for t in passed_list:
        ws_passed.append([t['id'], t['category'], t['name'], round(t.get('duration', 0.0), 3), t['priority']])
        style_row(ws_passed, ws_passed.max_row, border=thin_border)
    auto_fit_columns(ws_passed)

    # Sheet 3: Failed Tests
    ws_failed = wb.create_sheet(title="Failed Tests")
    ws_failed.append(["Test ID", "Module", "Test Name", "Failure Reason", "Execution Time (s)"])
    style_row(ws_failed, 1, font=font_header, fill=fill_header, align=align_center)
    for t in failed_list:
        ws_failed.append([t['id'], t['category'], t['name'], t.get('failure_reason', ''), round(t.get('duration', 0.0), 3)])
        style_row(ws_failed, ws_failed.max_row, border=thin_border)
    auto_fit_columns(ws_failed)

    # Sheet 4: Skipped Tests
    ws_skipped = wb.create_sheet(title="Skipped Tests")
    ws_skipped.append(["Test ID", "Module", "Test Name", "Priority"])
    style_row(ws_skipped, 1, font=font_header, fill=fill_header, align=align_center)
    for t in skipped_list:
        ws_skipped.append([t['id'], t['category'], t['name'], t['priority']])
        style_row(ws_skipped, ws_skipped.max_row, border=thin_border)
    auto_fit_columns(ws_skipped)

    # Sheet 5: Execution Metrics
    ws_metrics = wb.create_sheet(title="Execution Metrics")
    ws_metrics.append(["Metric Parameter", "Value"])
    style_row(ws_metrics, 1, font=font_header, fill=fill_header, align=align_center)
    metrics_data = [
        ["Total Test Cases Generated", total_tests],
        ["Total Executed", passed_count + failed_count],
        ["Passed Tests", passed_count],
        ["Failed Tests", failed_count],
        ["Skipped Tests", skipped_count],
        ["Success Rate (%)", f"{success_rate:.2f}%"],
        ["Total Duration (s)", f"{total_duration:.3f} s"]
    ]
    for row in metrics_data:
        ws_metrics.append(row)
        style_row(ws_metrics, ws_metrics.max_row, border=thin_border)
    auto_fit_columns(ws_metrics)

    # Sheet 6: Defect Summary
    ws_defect = wb.create_sheet(title="Defect Summary")
    ws_defect.append(["Test ID", "Module", "Test Name", "Failure Diagnosis / Stacktrace"])
    style_row(ws_defect, 1, font=font_header, fill=fill_header, align=align_center)
    for t in failed_list:
        ws_defect.append([t['id'], t['category'], t['name'], t.get('failure_reason', 'Assertion Failure')])
        style_row(ws_defect, ws_defect.max_row, border=thin_border)
    auto_fit_columns(ws_defect)

    wb.save(os.path.join(settings.EXCEL_DIR, 'Automation_Test_Report.xlsx'))

    # 2. Failed_Test_Cases.xlsx
    wb_failed = Workbook()
    ws = wb_failed.active
    ws.title = "Failed Test Cases"
    ws.append(["Test ID", "Module", "Test Name", "Failure Reason", "Priority"])
    style_row(ws, 1, font=font_header, fill=PatternFill(start_color='FF3333', end_color='FF3333', fill_type='solid'), align=align_center)
    for t in failed_list:
        ws.append([t['id'], t['category'], t['name'], t.get('failure_reason', ''), t['priority']])
        style_row(ws, ws.max_row, border=thin_border)
    auto_fit_columns(ws)
    wb_failed.save(os.path.join(settings.EXCEL_DIR, 'Failed_Test_Cases.xlsx'))

    # 3. Passed_Test_Cases.xlsx
    wb_passed = Workbook()
    ws = wb_passed.active
    ws.title = "Passed Test Cases"
    ws.append(["Test ID", "Module", "Test Name", "Priority", "Duration (s)"])
    style_row(ws, 1, font=font_header, fill=PatternFill(start_color='10A866', end_color='10A866', fill_type='solid'), align=align_center)
    for t in passed_list:
        ws.append([t['id'], t['category'], t['name'], t['priority'], round(t.get('duration', 0.0), 3)])
        style_row(ws, ws.max_row, border=thin_border)
    auto_fit_columns(ws)
    wb_passed.save(os.path.join(settings.EXCEL_DIR, 'Passed_Test_Cases.xlsx'))

    # 4. Summary_Report.xlsx
    wb_summary = Workbook()
    ws = wb_summary.active
    ws.title = "Summary Report"
    ws.append(["Summary Metric", "Value"])
    style_row(ws, 1, font=font_header, fill=fill_header, align=align_center)
    for row in metrics_data:
        ws.append(row)
        style_row(ws, ws.max_row, border=thin_border)
    auto_fit_columns(ws)
    wb_summary.save(os.path.join(settings.EXCEL_DIR, 'Summary_Report.xlsx'))
