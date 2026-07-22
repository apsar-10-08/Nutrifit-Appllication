import os
import shutil
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

os.makedirs("testing-reports", exist_ok=True)

# 35 Discovered Screens / Modules from lib/main.dart
SCREENS_AND_MODULES = [
    ("SplashScreen", "Splash & Onboarding Screen"),
    ("LoginScreen", "Login & Authentication Screen"),
    ("SignUpScreen", "User Registration Screen"),
    ("ForgotScreen", "Forgot Password Recovery Screen"),
    ("SelectScreen", "Onboarding Primary Goal Screen"),
    ("GoalScreen", "Fitness Target Goal Screen"),
    ("GenderScreen", "Demographics & Gender Screen"),
    ("FoodScreen", "Dietary Preference & Food Screen"),
    ("LocationScreen", "User Location & Region Screen"),
    ("NumberScreen", "Body Metrics & Measurements Screen"),
    ("DashboardScreen", "Main Dashboard & Navigation Shell"),
    ("HomeTab", "Home Overview & Daily Progress Tab"),
    ("HeroCard", "Today Workout & Hero Banner Section"),
    ("WorkoutDayCard", "Daily Exercise Routine & Set Tracking"),
    ("WarmupCard", "Stretching & Warm-up Routine View"),
    ("DietDayCard", "Diet & Meal Planner View"),
    ("BudgetFoodSection", "Budget Diet & Low-Cost Meal Planner"),
    ("PlansTab", "Workout & Meal Plans Catalog Tab"),
    ("TrackersTab", "Health Trackers Overview Tab"),
    ("WaterTrackerSection", "Water Intake & Hydration Tracker"),
    ("SleepTrackerSection", "Sleep Efficiency & REM Tracker"),
    ("StepTrackerSection", "Step Counter & Distance Burn Tracker"),
    ("AITrainerScreen", "AI Fitness & Exercise Trainer Screen"),
    ("ShopTab", "Supplement & Equipment Shop Catalog Tab"),
    ("BudgetProductCard", "Budget Supplement & Product Details"),
    ("CartScreen", "Shopping Cart & Item Summary Screen"),
    ("CheckoutScreen", "Checkout & Payment Gateway Screen"),
    ("AddressManagementSection", "Shipping Address & Delivery Form"),
    ("OrderConfirmationScreen", "Order Confirmation & Receipt Screen"),
    ("MyOrdersScreen", "My Orders Stepper & History Screen"),
    ("OrderTrackingTimeline", "Live Order Tracking & Stepper Timeline"),
    ("ProfileTab", "User Profile & Account Overview Tab"),
    ("RemindersNotificationModule", "Reminders & Push Notifications Module"),
    ("SettingsLocalizationModule", "Language Switcher & Theme Settings"),
    ("SupabaseIntegrationModule", "Supabase PostgreSQL Database Sync")
]

COLUMNS = [
    "Test Case ID", "Module/Screen", "Test Category", "Test Scenario", "Test Steps",
    "Test Data", "Expected Result", "Actual Result", "Status", "Priority",
    "Severity", "Automation Status", "Environment", "Browser/Device", "Remarks"
]

WORKBOOK_CONFIGS = [
    ("NutriFit_Selenium_Test_Report.xlsx", "Selenium Web UI Testing", "SEL", "automation/pages/test_selenium_web.py", "Staging / Web Chrome", "Chrome v126 / Edge v126"),
    ("NutriFit_Appium_Test_Report.xlsx", "Appium Android E2E Testing", "APP", "test/e2e/run_appium_test.py", "Android 14 Emulator", "Android Pixel 6 (API 34)"),
    ("NutriFit_E2E_Test_Report.xlsx", "End-to-End Workflow Testing", "E2E", "test/e2e/test_full_workflow.py", "Integration Environment", "Cross-Platform Web & Android"),
    ("NutriFit_Functional_Test_Report.xlsx", "Functional & Integration Testing", "FUN", "test/integration_test.dart", "Unit / Flutter Test Suite", "Dart VM / Flutter SDK"),
    ("NutriFit_Load_Test_Report.xlsx", "Load & Performance Testing", "LOD", "test/load/k6_load_test.js", "Locust / k6 Load Environment", "Distributed Load Generator Nodes"),
    ("NutriFit_Vulnerability_Test_Report.xlsx", "Safe Vulnerability & Security Checks", "SEC", "test/e2e/generate_vulnerability_report.py", "Security Audit Sandbox", "Defensive Security Audit Engine")
]

CATEGORY_SCENARIOS = {
    "SEL": [
        ("Positive Web UI Layout", "Launch browser, navigate to {screen}, verify grid container alignment and responsive element scaling.", "Viewport: 1920x1080, Screen: {screen}", "UI components align cleanly without scrollbar clipping or overflow."),
        ("Negative Input Form Validation", "Focus input controls on {screen}, enter out-of-bound strings, trigger blur event.", "Input: invalid_string, Screen: {screen}", "Validation badge displays accessible aria error message."),
        ("Boundary Viewport Responsiveness", "Resize web browser window from 320px mobile breakpoint up to 4K desktop width.", "Breakpoint Widths: [320, 768, 1024, 1440, 3840]", "Layout adapts responsively without overlapping UI text nodes."),
        ("Interactive Hover & Focus Traversal", "Tab through interactive focusable nodes on {screen}, hover cursor over CTA buttons.", "Key Navigation: TAB, Hover target: {screen}", "Focus ring highlights elements in logical DOM tab order."),
        ("Cross-Browser Rendering Parity", "Execute UI test suite for {screen} across Chrome, Firefox, Safari, and Edge.", "Browsers: Chrome, Firefox, Safari, Edge", "DOM element styles render identically with zero CSS misalignment."),
        ("Arabic Right-to-Left (RTL) Layout", "Switch locale to Arabic (ar_SA) on {screen}, inspect text direction and icon mirroring.", "Locale: ar_SA, Target: {screen}", "UI flips to right-to-left layout with mirrored icons and aligned text."),
        ("Localization & Multi-Language Switching", "Toggle header language dropdown on {screen} between English, Hindi, and Arabic.", "Locales: en, hi, ar", "All visible text strings update to selected language instantly."),
        ("Data Table & Grid State Retention", "Interact with data filters or state controls on {screen}, navigate away and return.", "State Payload: {screen}_filter", "Screen preserves applied filter state without unexpected reset."),
        ("Modal Backdrop & Keyboard Escape", "Trigger pop-up dialog/modal on {screen}, press Escape key to dismiss.", "Modal Target: {screen}", "Modal closes smoothly and returns focus to parent container."),
        ("Network Failure UI Fallback", "Throttled network to offline mode while viewing {screen}, attempt UI interaction.", "Network: Offline, Target: {screen}", "User-friendly offline banner appears with retry button.")
    ],
    "APP": [
        ("Native Touch Tap & Ripple Effect", "Launch Android emulator, navigate to {screen}, tap primary interactive buttons.", "Touch Target: {screen}_btn", "Native Material ripple effect triggers with zero tap delay."),
        ("Vertical Swipe & Scroll Velocity", "Perform vertical swipe gesture on scrollable list container within {screen}.", "Gesture: Swipe Up/Down, Speed: 1500px/s", "RecyclerView/ListView scrolls smoothly at 60fps without lag."),
        ("Screen Auto-Rotation (Portrait/Landscape)", "Rotate device orientation from portrait to landscape 90deg on {screen}.", "Orientation: Landscape 90deg", "Layout recalculates safe bounds in landscape mode without clipping."),
        ("Hardware Back Button Navigation", "Navigate deep into {screen}, press physical/software Android Back button (KeyCode 4).", "KeyCode: 4 (BACK)", "App pops top route cleanly and returns to previous screen."),
        ("System Permission Dialog Handling", "Trigger feature requiring permission on {screen}, handle native system dialog.", "Permission: Camera/Storage/Biometrics", "System permission dialog handles Grant/Deny gracefully."),
        ("Push Notification Shade & Routing", "Dispatch local push notification alert for {screen}, tap banner in shade.", "Notification ID: {screen}_push", "App resumes from background and routes directly to target screen."),
        ("Offline Flight Mode & Hive Cache", "Enable Airplane Mode on device, navigate to {screen}, verify cached data display.", "Network: Flight Mode", "Screen renders locally cached Hive storage data with offline indicator."),
        ("Biometric Authentication Lock", "Trigger security action on {screen} requiring BiometricPrompt fingerprint/face ID.", "Auth Payload: Biometrics", "Native Android BiometricPrompt appears; success unlocks view."),
        ("App Backgrounding & Memory Resume", "Press Home button on {screen}, wait 30 seconds in background, re-open app.", "Lifecycle: OnPause -> OnResume", "App restores exact scroll position and state without crashing."),
        ("Multi-Window & Split Screen Scaling", "Drag app into Android Multi-Window split screen mode while viewing {screen}.", "Window Mode: Split Screen 50/50", "Screen resizes UI components responsively within half-height viewport.")
    ],
    "E2E": [
        ("End-to-End User Journey Entry", "Initiate multi-step user journey starting from {screen}, complete initial steps.", "Journey Step 1: {screen}", "Step 1 completes successfully and transitions to next workflow step."),
        ("Cross-Screen Data Mutation Sync", "Update data records on {screen}, navigate to secondary screen, inspect state.", "Mutation Payload: {screen}_sync", "Secondary screen reflects updated state immediately without refresh."),
        ("Wizard Form Forward Navigation", "Fill all required inputs on {screen}, click Next button to advance workflow.", "Wizard Inputs: {screen}_valid", "Form validates inputs and transitions cleanly to next wizard step."),
        ("Wizard Form Back-Tracking & Retention", "Advance to step 3, click Back button returning to {screen}, inspect inputs.", "Wizard State: Step 3 -> {screen}", "Previously entered form values remain preserved in input fields."),
        ("Database Transactional Sync", "Submit transaction on {screen}, verify client cache and Supabase DB sync.", "DB Transaction: {screen}_submit", "Record persists in local cache and syncs to Supabase PostgreSQL table."),
        ("Error Boundary Recovery & Rollback", "Simulate network error during transaction on {screen}, observe error boundary.", "Fault Injection: 500 Server Error", "Transaction rolls back safely; error toast displays retry CTA."),
        ("Multi-Tab Session Synchronization", "Open secondary app view on {screen}, update user state, verify session sync.", "Session Token: {screen}_session", "User session and state stay synchronized across active views."),
        ("Payment Gateway & Receipt Flow", "Proceed through checkout workflow involving {screen}, complete transaction.", "Transaction Payload: {screen}_order", "Payment gateway returns success token; confirmation receipt renders."),
        ("Order Stepper Status Progress", "Observe live status updates on {screen} as order progresses through steps.", "Order ID: {screen}_ord_101", "Status stepper updates sequentially from Placed to Delivered."),
        ("Full Profile & Settings Transition", "Navigate from {screen} to profile settings, update preferences, return to screen.", "Profile Prefs: {screen}_locale", "Screen re-renders with updated user preferences applied globally.")
    ],
    "FUN": [
        ("Feature Input Validation Rules", "Pass boundary, empty, null, and special character strings to {screen} input handler.", "Input: {screen}_edge_case", "Validation handler enforces rules and returns clear error strings."),
        ("Business Logic Calculation Math", "Invoke business logic calculation engine on {screen} with test dataset.", "Dataset: {screen}_calc_data", "Engine calculates precise values according to domain business rules."),
        ("Async Service Call & JSON Parsing", "Execute asynchronous service call on {screen}, parse response domain model.", "API Payload: {screen}_json", "JSON payload parses cleanly into strongly typed Dart model class."),
        ("Reactive State Management Rebuild", "Mutate state property bound to {screen}, count widget rebuild notifications.", "State Provider: {screen}_notify", "Widget list listener triggers single UI rebuild with updated state."),
        ("Role Authorization Guard Check", "Attempt accessing feature on {screen} with restricted user permission role.", "User Role: Restricted_Guest", "Authorization guard blocks access and displays permission error."),
        ("Data Persistence & Storage Invalidation", "Save record on {screen}, trigger cache invalidation, reload screen view.", "Cache Key: {screen}_key", "Old cache clears cleanly; screen fetches fresh record from storage."),
        ("Formula Range & Unit Conversion", "Input metric values on {screen}, verify imperial/metric unit conversion math.", "Unit Input: {screen}_units", "Converts units accurately (kg to lbs, km to miles) without rounding drift."),
        ("State Machine Allowed Transitions", "Attempt state transition on {screen} (e.g., Pending -> Completed).", "Transition: {screen}_state", "State machine permits valid transitions while blocking illegal skips."),
        ("UI Event Dispatch & Callback Execution", "Trigger interactive event callback on {screen}, verify event arguments.", "Event Spec: {screen}_evt", "Callback dispatches with correct event parameter arguments."),
        ("Memory Heap Release & Disposal", "Push {screen} onto navigation stack, pop screen, verify controller disposal.", "Lifecycle: Dispose Controller", "Controllers and listeners dispose cleanly without memory leaks.")
    ],
    "LOD": [
        ("50 Virtual Users Concurrent Baseline", "Simulate 50 concurrent virtual users interacting with {screen} endpoints.", "VUs: 50, Hatch Rate: 5/s", "P95 latency remains under 200ms with zero HTTP errors."),
        ("100 Virtual Users Concurrency Peak", "Scale concurrent virtual user load to 100 VUs targeting {screen} API.", "VUs: 100, Duration: 3m", "P95 response latency stays below 300ms SLA limit."),
        ("250 Virtual Users High Load Surge", "Inject 250 concurrent user requests targeting {screen} data endpoints.", "VUs: 250, Ramp: 10s", "System throughput scales linearly without API gateway drops."),
        ("500 Virtual Users Heavy Stress Test", "Push load to 500 VUs on {screen} to evaluate connection pool limits.", "VUs: 500, Duration: 5m", "Database connection pool maintains stability with 0.0% failure rate."),
        ("1000 Virtual Users Maximum Saturation", "Run 1000 concurrent VUs on {screen} to determine maximum saturation limit.", "VUs: 1000, Saturation Test", "System gracefully degrades response time without 5xx crashes."),
        ("Spike Surge 10x Burst Ingestion", "Trigger sudden 10x traffic spike on {screen} within a 5-second window.", "Spike: 10x baseline VUs", "API gateway queues burst traffic and recovers baseline latency."),
        ("Endurance Sustained Load 1-Hour", "Sustain 80% maximum capacity load on {screen} over 1-hour test run.", "Duration: 60 minutes", "Memory and CPU telemetry remain stable without memory leaks."),
        ("Database Lock Contention Benchmark", "Execute 200 concurrent write transactions on {screen} database tables.", "Writes: 200 concurrent", "Database handles optimistic locks without primary key deadlocks."),
        ("Realtime WebSocket Subscription Load", "Open 300 active WebSocket channels receiving live updates for {screen}.", "WebSockets: 300 channels", "Realtime server broadcasts payload updates with latency < 50ms."),
        ("Asset Loading & CDN Hit Speed", "Load media feed assets associated with {screen} under throttled network.", "Assets: 50 WebP files", "Cache hit speed < 30ms; CDN serves media assets efficiently.")
    ],
    "SEC": [
        ("Authentication Lockout & Brute-Force", "Attempt 10 invalid authentication requests on {screen} in rapid succession.", "Attempts: 10 invalid", "Account temporarily locks out after 5 failed attempts with 429 response."),
        ("Role-Based Authorization & IDOR Check", "Attempt querying secondary user private records on {screen} using guest JWT.", "IDOR Payload: {screen}_user_999", "Database and API reject request with HTTP 403 Forbidden."),
        ("Session Token Expiration & Revocation", "Present expired JWT token to {screen} API endpoint, attempt data fetch.", "JWT Token: Expired/Revoked", "API gateway invalidates session and redirects to login view."),
        ("Safe SQL Injection Input Defense", "Submit safe SQLi payload `' OR '1'='1` into text input fields on {screen}.", "Payload: `' OR '1'='1`", "Input is parameterized/escaped cleanly; no SQL error occurs."),
        ("Safe XSS Script Escaping Validation", "Submit safe script payload `<script>alert(1)</script>` into {screen} fields.", "Payload: `<script>alert(1)</script>`", "Payload renders as plain text string; script does not execute."),
        ("TLS 1.3 Transport Encryption Enforcement", "Inspect outbound network traffic for {screen} using passive OWASP proxy.", "Protocol: TLS 1.3 / HTTPS", "All HTTP connections upgrade to TLS 1.3; unencrypted HTTP blocked."),
        ("API Rate Limiting 429 Response Check", "Send 100 rapid requests within 5 seconds to {screen} public endpoints.", "Requests: 100 in 5s", "Rate limiter triggers HTTP 429 Too Many Requests response."),
        ("Supabase Row Level Security (RLS) Audit", "Query Supabase PostgreSQL table for {screen} without auth header.", "RLS Check: {screen}_table", "Supabase RLS policy blocks unauthorized row query access."),
        ("Privileged Route Guard Verification", "Attempt deep linking directly to restricted {screen} view as unauthenticated user.", "Route: /{screen}_admin", "Router guard intercepts request and redirects to login screen."),
        ("Sensitive Data Masking & Log Leak Check", "Audit application log files and network payloads for exposed {screen} secrets.", "Audit: Password/Token fields", "Sensitive credentials and PII are masked or omitted from logs.")
    ]
}

def generate_10_cases_per_screen(cat_code, cat_title, default_script_path, default_env, default_dev):
    cases = []
    tc_counter = 1
    scenarios_tpl = CATEGORY_SCENARIOS[cat_code]

    for scr_code, scr_title in SCREENS_AND_MODULES:
        for idx in range(10):
            tc_id = f"TC-{tc_counter:03d}"
            scen_tpl = scenarios_tpl[idx]

            scen_title = scen_tpl[0].format(screen=scr_code)
            steps = f"1. Launch test runner in {default_env}.\n2. Navigate to {scr_title} ({scr_code}).\n3. {scen_tpl[1].format(screen=scr_code)}\n4. Verify expected results and log evidence."
            test_data = scen_tpl[2].format(screen=scr_code)
            expected = scen_tpl[3].format(screen=scr_code)

            prio = "High" if (tc_counter % 3 == 0) else ("Low" if (tc_counter % 5 == 0) else "Medium")
            sev = "Critical" if (tc_counter % 7 == 0) else ("Major" if (tc_counter % 2 == 0) else "Moderate")

            cases.append({
                "Test Case ID": tc_id,
                "Module/Screen": scr_code,
                "Test Category": cat_title,
                "Test Scenario": f"{scen_title} on {scr_title}",
                "Test Steps": steps,
                "Test Data": test_data,
                "Expected Result": expected,
                "Actual Result": "Pending Execution",
                "Status": "Not Executed",
                "Priority": prio,
                "Severity": sev,
                "Automation Status": "Script Ready",
                "Environment": default_env,
                "Browser/Device": default_dev,
                "Remarks": f"Screen-specific baseline case for {scr_title}"
            })
            tc_counter += 1

    return cases

def build_workbook(filename, cat_title, cases):
    wb = openpyxl.Workbook()
    wb.remove(wb.active) # remove default sheet

    font_family = "Segoe UI"
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="10A866", end_color="10A866", fill_type="solid") # Emerald Green

    title_font = Font(name=font_family, size=15, bold=True, color="0A5C36")
    subtitle_font = Font(name=font_family, size=10, italic=True, color="555555")
    section_font = Font(name=font_family, size=12, bold=True, color="0A5C36")

    thin_border = Border(
        left=Side(style='thin', color='E0E0E0'),
        right=Side(style='thin', color='E0E0E0'),
        top=Side(style='thin', color='E0E0E0'),
        bottom=Side(style='thin', color='E0E0E0')
    )

    card_border = Border(
        left=Side(style='medium', color='10A866'),
        right=Side(style='medium', color='10A866'),
        top=Side(style='medium', color='10A866'),
        bottom=Side(style='medium', color='10A866')
    )

    status_styles = {
        "Passed": (PatternFill(start_color="E8F8F0", end_color="E8F8F0", fill_type="solid"), Font(name=font_family, color="10A866", bold=True)),
        "Failed": (PatternFill(start_color="FCE8E6", end_color="FCE8E6", fill_type="solid"), Font(name=font_family, color="D93025", bold=True)),
        "Skipped": (PatternFill(start_color="FEF7E0", end_color="FEF7E0", fill_type="solid"), Font(name=font_family, color="B06000", bold=True)),
        "Blocked": (PatternFill(start_color="FFEFE6", end_color="FFEFE6", fill_type="solid"), Font(name=font_family, color="D9530F", bold=True)),
        "Not Executed": (PatternFill(start_color="F1F3F4", end_color="F1F3F4", fill_type="solid"), Font(name=font_family, color="5F6368", bold=True))
    }

    total_screen_count = len(SCREENS_AND_MODULES)
    total_case_count = len(cases)

    # ----------------------------------------------------
    # Sheet 1: Summary Dashboard
    # ----------------------------------------------------
    ws_sum = wb.create_sheet(title="Summary Dashboard")
    ws_sum.views.sheetView[0].showGridLines = True

    ws_sum["A1"] = f"NutriFit QA Report - {cat_title}"
    ws_sum["A1"].font = title_font
    ws_sum["A2"] = f"Report File: {filename} | Total Screens/Modules: {total_screen_count} | 10 Cases Per Screen | Total: {total_case_count} Cases"
    ws_sum["A2"].font = subtitle_font

    ws_sum.row_dimensions[1].height = 24
    ws_sum.row_dimensions[2].height = 18

    # KPI Summary Cards
    kpis = [
        ("Total Test Cases", "=COUNT('Detailed Test Cases'!A:A)"),
        ("Passed", '=COUNTIF(\'Detailed Test Cases\'!I:I, "Passed")'),
        ("Failed", '=COUNTIF(\'Detailed Test Cases\'!I:I, "Failed")'),
        ("Skipped", '=COUNTIF(\'Detailed Test Cases\'!I:I, "Skipped")'),
        ("Blocked", '=COUNTIF(\'Detailed Test Cases\'!I:I, "Blocked")'),
        ("Not Executed", '=COUNTIF(\'Detailed Test Cases\'!I:I, "Not Executed")'),
        ("Pass Rate %", '=IF(B4=0, "0%", TEXT(B5/B4, "0.0%"))')
    ]

    for col_idx, (hdr, form) in enumerate(kpis, start=1):
        c_hdr = ws_sum.cell(row=4, column=col_idx, value=hdr)
        c_hdr.font = Font(name=font_family, size=9, bold=True, color="555555")
        c_hdr.alignment = Alignment(horizontal='center', vertical='center')
        c_hdr.fill = PatternFill(start_color="F5F7FA", end_color="F5F7FA", fill_type="solid")

        c_val = ws_sum.cell(row=5, column=col_idx, value=form)
        c_val.font = Font(name=font_family, size=14, bold=True, color="0A5C36")
        c_val.alignment = Alignment(horizontal='center', vertical='center')
        c_val.border = card_border

    # Screen Coverage Breakdown Table
    ws_sum["A7"] = f"Screen/Module-Wise Test Coverage Breakdown (All {total_screen_count} Screens)"
    ws_sum["A7"].font = section_font

    mod_headers = ["Module/Screen", "Total Cases", "Passed", "Failed", "Skipped", "Blocked", "Not Executed"]
    ws_sum.append(mod_headers)
    ws_sum.row_dimensions[8].height = 24

    for c_idx in range(1, len(mod_headers) + 1):
        cell = ws_sum.cell(row=8, column=c_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for r_offset, (scr_code, scr_title) in enumerate(SCREENS_AND_MODULES, start=9):
        ws_sum.cell(row=r_offset, column=1, value=scr_code)
        ws_sum.cell(row=r_offset, column=2, value=f'=COUNTIF(\'Detailed Test Cases\'!B:B, A{r_offset})')
        ws_sum.cell(row=r_offset, column=3, value=f'=COUNTIFS(\'Detailed Test Cases\'!B:B, A{r_offset}, \'Detailed Test Cases\'!I:I, "Passed")')
        ws_sum.cell(row=r_offset, column=4, value=f'=COUNTIFS(\'Detailed Test Cases\'!B:B, A{r_offset}, \'Detailed Test Cases\'!I:I, "Failed")')
        ws_sum.cell(row=r_offset, column=5, value=f'=COUNTIFS(\'Detailed Test Cases\'!B:B, A{r_offset}, \'Detailed Test Cases\'!I:I, "Skipped")')
        ws_sum.cell(row=r_offset, column=6, value=f'=COUNTIFS(\'Detailed Test Cases\'!B:B, A{r_offset}, \'Detailed Test Cases\'!I:I, "Blocked")')
        ws_sum.cell(row=r_offset, column=7, value=f'=COUNTIFS(\'Detailed Test Cases\'!B:B, A{r_offset}, \'Detailed Test Cases\'!I:I, "Not Executed")')

        for c_idx in range(1, 8):
            cell = ws_sum.cell(row=r_offset, column=c_idx)
            cell.font = Font(name=font_family, size=9)
            cell.border = thin_border
            if c_idx >= 2:
                cell.alignment = Alignment(horizontal='center', vertical='center')

    # Total Row
    tot_row = 9 + len(SCREENS_AND_MODULES)
    ws_sum.cell(row=tot_row, column=1, value="Total Suite").font = Font(name=font_family, size=10, bold=True)
    for c_idx in range(2, 8):
        col_let = get_column_letter(c_idx)
        cell = ws_sum.cell(row=tot_row, column=c_idx, value=f"=SUM({col_let}9:{col_let}{tot_row-1})")
        cell.font = Font(name=font_family, size=10, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        cell.fill = PatternFill(start_color="F0F9F4", end_color="F0F9F4", fill_type="solid")

    ws_sum.column_dimensions['A'].width = 32
    for c in ['B', 'C', 'D', 'E', 'F', 'G']:
        ws_sum.column_dimensions[c].width = 16

    # ----------------------------------------------------
    # Sheet 2: Detailed Test Cases
    # ----------------------------------------------------
    ws_det = wb.create_sheet(title="Detailed Test Cases")
    ws_det.views.sheetView[0].showGridLines = True
    ws_det.append(COLUMNS)
    ws_det.row_dimensions[1].height = 28
    ws_det.freeze_panes = "A2"

    for c_idx in range(1, len(COLUMNS) + 1):
        cell = ws_det.cell(row=1, column=c_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for r_idx, tc in enumerate(cases, start=2):
        row_vals = [tc[c] for c in COLUMNS]
        ws_det.append(row_vals)
        ws_det.row_dimensions[r_idx].height = 36

        for c_idx in range(1, len(COLUMNS) + 1):
            cell = ws_det.cell(row=r_idx, column=c_idx)
            cell.font = Font(name=font_family, size=9)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='top', wrap_text=True)

            if c_idx in [1, 9, 10, 11, 12]: # ID, Status, Priority, Severity, Automation Status
                cell.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)

            # Apply Status Style (9th column: Status)
            if c_idx == 9:
                status_val = cell.value
                if status_val in status_styles:
                    fill, font = status_styles[status_val]
                    cell.fill = fill
                    cell.font = font

    ws_det.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{len(cases) + 1}"

    col_widths = {
        "Test Case ID": 14,
        "Module/Screen": 28,
        "Test Category": 25,
        "Test Scenario": 35,
        "Test Steps": 45,
        "Test Data": 30,
        "Expected Result": 35,
        "Actual Result": 20,
        "Status": 15,
        "Priority": 12,
        "Severity": 12,
        "Automation Status": 18,
        "Environment": 25,
        "Browser/Device": 25,
        "Remarks": 25
    }

    for c_idx, c_name in enumerate(COLUMNS, start=1):
        ws_det.column_dimensions[get_column_letter(c_idx)].width = col_widths.get(c_name, 20)

    # ----------------------------------------------------
    # Sheet 3: Execution Evidence
    # ----------------------------------------------------
    ws_ev = wb.create_sheet(title="Execution Evidence")
    ws_ev.views.sheetView[0].showGridLines = True
    ev_cols = ["Test Case ID", "Module/Screen", "Test Scenario", "Status", "Evidence Path", "Timestamp", "Verification Notes"]
    ws_ev.append(ev_cols)
    ws_ev.row_dimensions[1].height = 28
    ws_ev.freeze_panes = "A2"

    for c_idx in range(1, len(ev_cols) + 1):
        cell = ws_ev.cell(row=1, column=c_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for idx, tc in enumerate(cases[:20], start=2):
        ws_ev.append([
            tc["Test Case ID"],
            tc["Module/Screen"],
            tc["Test Scenario"],
            tc["Status"],
            f"testing-reports/evidence/{tc['Test Case ID'].lower()}_proof.png",
            "2026-07-22 00:00:00 UTC",
            "Baseline execution artifact initialized"
        ])
        for c_idx in range(1, len(ev_cols) + 1):
            cell = ws_ev.cell(row=idx, column=c_idx)
            cell.font = Font(name=font_family, size=9)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='top', wrap_text=True)

    ws_ev.auto_filter.ref = f"A1:{get_column_letter(len(ev_cols))}21"
    for c_idx in range(1, len(ev_cols) + 1):
        ws_ev.column_dimensions[get_column_letter(c_idx)].width = 25

    # Save to testing-reports/ AND root directory
    rep_path = os.path.join("testing-reports", filename)
    wb.save(rep_path)
    shutil.copy(rep_path, filename)
    print(f"SUCCESS: Saved '{rep_path}' and '{filename}' with {total_case_count} test cases.")

def main():
    total_cases_all = 0
    num_screens = len(SCREENS_AND_MODULES)
    print(f"=== GENERATING WORKBOOKS FOR {num_screens} DISCOVERED SCREENS/MODULES (10 CASES EACH) ===")

    for filename, cat_title, prefix, script_path, env, dev in WORKBOOK_CONFIGS:
        cases = generate_10_cases_per_screen(prefix, cat_title, script_path, env, dev)
        expected_cases = num_screens * 10
        assert len(cases) == expected_cases, f"Error: expected {expected_cases} cases for {filename}, got {len(cases)}"
        build_workbook(filename, cat_title, cases)
        total_cases_all += len(cases)

    print("\n==================================================")
    print(f"GENERATION COMPLETE! TOTAL TEST CASES = {total_cases_all}")
    print(f"FILES GENERATED = {len(WORKBOOK_CONFIGS)}")
    print("==================================================")

if __name__ == "__main__":
    main()
