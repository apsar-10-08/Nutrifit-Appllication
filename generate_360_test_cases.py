import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Ensure testing-reports directory exists
os.makedirs("testing-reports", exist_ok=True)

# ----------------------------------------------------
# Define 37 Required Modules
# ----------------------------------------------------
MODULES = [
    "Splash", "Login", "Signup", "Forgot Password", "Email Verification",
    "Onboarding", "Dashboard", "Profile", "Language Settings", "AI Trainer",
    "Workout Plan", "Diet Plan", "Budget Diet Plan", "Stretching and Warm-up",
    "Water Tracker", "Sleep Tracker", "Step Tracker", "Calorie Calculator",
    "Meal Reminder", "Notifications", "Shop", "Product Search and Filters",
    "Product Details", "Wishlist", "Cart", "Checkout", "Address Management",
    "Payment Method", "Place Order", "My Orders", "Order Details",
    "Order Cancellation", "Supabase Integration", "Responsive Web UI",
    "Android UI", "Error Handling", "Logout"
]

TEST_TYPES = [
    "Selenium Web UI Testing",
    "Appium Android E2E Testing",
    "Flutter Unit and Widget Testing",
    "Functional and Integration Testing",
    "Load and Performance Testing",
    "Safe Vulnerability and Configuration Checks"
]

COLUMNS = [
    "Test Case ID", "Testing Type", "Module", "Test Scenario", "Test Case Title",
    "Objective", "Preconditions", "Test Steps", "Test Data", "Expected Result",
    "Actual Result", "Priority", "Severity", "Status", "Evidence Path", "Remarks"
]

# Helper function to generate structured test cases
def generate_all_360_test_cases():
    test_cases = []

    # Helper to add test case
    def add_tc(tc_id, t_type, module, scenario, title, obj, pre, steps, data, exp, prio="Medium", sev="Moderate"):
        test_cases.append({
            "Test Case ID": tc_id,
            "Testing Type": t_type,
            "Module": module,
            "Test Scenario": scenario,
            "Test Case Title": title,
            "Objective": obj,
            "Preconditions": pre,
            "Test Steps": steps,
            "Test Data": data,
            "Expected Result": exp,
            "Actual Result": "Pending Execution",
            "Priority": prio,
            "Severity": sev,
            "Status": "Not Executed",
            "Evidence Path": f"evidence/{tc_id.lower()}_screenshot.png",
            "Remarks": "Automated baseline generation"
        })

    # =========================================================================
    # 1. Selenium Web UI Testing (100 Test Cases: SEL-WEB-001 to SEL-WEB-100)
    # =========================================================================
    selenium_modules = (MODULES * 3)[:100] # Ensure all modules covered across 100 cases
    
    web_scenarios = [
        ("Verify page load responsiveness across desktop break points", "Launch web browser at 1920x1080 resolution, navigate to portal, verify layout grid and header alignment.", "Standard viewport 1920x1080", "Layout scales seamlessly without horizontal overflow or overlapping text elements."),
        ("Verify form input validations and error hints", "Focus input field, enter invalid format data, trigger blur event, inspect validation error badge.", "Invalid test string / out-of-bounds input", "Clear validation message displayed immediately under field."),
        ("Verify interactive button states and hover effects", "Hover mouse over primary action button, verify CSS transition, click button.", "Primary CTA button selector", "Button transforms smoothly on hover and handles click event correctly."),
        ("Verify modal dialog overlay behavior", "Trigger action opening dynamic modal, verify backdrop blur, click close icon.", "Modal trigger event", "Modal opens centered with focus trap; closes cleanly on overlay/icon click."),
        ("Verify table sorting and pagination controls", "Navigate to list view, click table column header for sorting, click next page button.", "Page size = 10", "Table updates sorted order instantly; page 2 records load cleanly."),
        ("Verify theme switching (Dark/Light mode) on Web", "Click theme toggle switch in global header navbar.", "Theme toggle selector", "UI colors switch smoothly across all visible DOM containers."),
        ("Verify navigation drawer opening and item selection", "Click hamburger menu icon, select target sidebar menu item.", "Sidebar nav items", "Drawer slides out smoothly and routes user to target page."),
        ("Verify image lazy loading and fallback rendering", "Scroll down page rapidly to trigger image viewports.", "High-res recipe / product images", "Images display lazy loading placeholders before rendering final image asset cleanly."),
        ("Verify browser back and forward navigation history", "Navigate from Dashboard -> Shop -> Product Details, click browser Back, then Forward.", "Browser history stack", "URL state and page DOM sync perfectly with browser history state."),
        ("Verify accessibility aria-labels and keyboard tab traversal", "Press TAB key repeatedly from top of web page to traverse focusable elements.", "Keyboard navigation", "Focus outline highlights each interactive element in logical DOM sequence.")
    ]

    for i in range(1, 101):
        tc_id = f"SEL-WEB-{i:03d}"
        mod = selenium_modules[i-1]
        scen_idx = (i - 1) % len(web_scenarios)
        template = web_scenarios[scen_idx]
        
        title = f"Web UI - {mod} - {template[0]} (#{i})"
        scenario = f"{template[0]} in {mod} module"
        obj = f"Ensure {mod} web UI component operates seamlessly under Chrome/Firefox/Edge Selenium driver."
        pre = f"Browser session active, navigated to NutriFit Web app {mod} route."
        steps = f"1. Open NutriFit Web app on Chrome/Edge.\n2. Navigate to {mod} screen.\n3. {template[1]}\n4. Capture DOM state and screenshot."
        data = f"Module: {mod}, Viewport: 1920x1080, Data: {template[2]}"
        exp = f"{template[3]} on {mod} module."
        prio = "High" if i % 3 == 0 else ("Low" if i % 5 == 0 else "Medium")
        sev = "Critical" if i % 7 == 0 else ("Major" if i % 2 == 0 else "Moderate")

        add_tc(tc_id, "Selenium Web UI Testing", mod, scenario, title, obj, pre, steps, data, exp, prio, sev)

    # =========================================================================
    # 2. Appium Android E2E Testing (90 Test Cases: APP-AND-001 to APP-AND-090)
    # =========================================================================
    appium_modules = (MODULES * 3)[:90]

    appium_scenarios = [
        ("Verify native touch swipe and scroll gestures", "Perform vertical swipe gesture on scrollable listview container.", "Android TouchAction / W3C Actions", "Container scrolls smoothly without frame drops or touch drag locking."),
        ("Verify hardware back button behavior", "Navigate deep into feature stack, press hardware back button.", "Android KeyCode 4 (BACK)", "App returns to previous screen cleanly without exiting app unexpectedly."),
        ("Verify device orientation toggle (Portrait to Landscape)", "Trigger device rotation to landscape, verify view adaptation.", "Device rotation 90deg", "UI resizes responsively in landscape without clipping key buttons."),
        ("Verify system notification banner pop-up and tap action", "Trigger local push alert, pull down Android notification shade, tap notification.", "Local notification payload", "App opens target screen specified in notification payload."),
        ("Verify biometric authentication prompt (Fingerprint/Face)", "Trigger sensitive action requiring biometric authentication.", "Mock biometric key", "Native Android BiometricPrompt appears; success unlocks target view."),
        ("Verify camera permission request dialog and image capture", "Tap photo upload button, handle system permission dialog, capture photo.", "CAMERA permission", "Permission granted dialog works; captured image sets profile/meal avatar."),
        ("Verify deep link url execution on Android device", "Execute adb command to launch deep link URI target into app.", "nutrifit://app/shop/item123", "App launches directly into target product details view."),
        ("Verify app backgrounding and resume lifecycle state", "Press Home button to send app to background for 30s, then re-open app.", "Android Lifecycle pause/resume", "App resumes state intact without memory leak or session logout."),
        ("Verify dark theme integration with Android system settings", "Toggle system-wide Android Dark Theme switch from quick settings.", "Android OS Theme setting", "NutriFit automatically adapts to OS dark theme preference."),
        ("Verify offline cached state when switching Airplane mode", "Turn on Airplane Mode on device, navigate app screens.", "Offline network state", "App displays offline warning banner while serving cached data gracefully.")
    ]

    for i in range(1, 91):
        tc_id = f"APP-AND-{i:03d}"
        mod = appium_modules[i-1]
        scen_idx = (i - 1) % len(appium_scenarios)
        template = appium_scenarios[scen_idx]

        title = f"Android E2E - {mod} - {template[0]} (#{i})"
        scenario = f"{template[0]} for Android {mod}"
        obj = f"Verify native Android application behavior for {mod} using Appium automation."
        pre = f"NutriFit APK installed on Android Emulator/Device (API level 33+), Appium session running."
        steps = f"1. Launch Appium driver session.\n2. Open {mod} view via native element selector.\n3. {template[1]}\n4. Verify UI state."
        data = f"Device: Pixel 7 Pro (API 34), Package: com.nutrifit.app, Module: {mod}"
        exp = f"{template[3]} for {mod} screen on Android."
        prio = "High" if i % 2 == 0 else "Medium"
        sev = "Major" if i % 3 == 0 else "Moderate"

        add_tc(tc_id, "Appium Android E2E Testing", mod, scenario, title, obj, pre, steps, data, exp, prio, sev)

    # =========================================================================
    # 3. Flutter Unit and Widget Testing (60 Test Cases: FLU-TST-001 to FLU-TST-060)
    # =========================================================================
    flutter_modules = (MODULES * 2)[:60]

    flutter_scenarios = [
        ("Verify model JSON serialization & deserialization logic", "Pass JSON map payload to Model.fromJson(), verify fields, call model.toJson().", "Valid JSON fixture", "Model instantiates correctly; toJson matches expected map key values."),
        ("Verify form field validation regex functions", "Invoke validation utility with valid, boundary, and malformed inputs.", "Test strings / numbers", "Validation function returns null for valid input and error string for invalid."),
        ("Verify ChangeNotifier state mutation and notifyListeners()", "Call state mutation method on Provider class, observe listener callback count.", "Provider state call", "State updates as expected and notifyListeners() triggers UI rebuild once."),
        ("Verify Widget rendering in WidgetTester environment", "Pump widget into WidgetTester tree, search elements using find.byType / find.text.", "Widget fixture context", "Widget renders without layout overflow warnings or unhandled exceptions."),
        ("Verify mock service dependency injection and async response handling", "Inject MockSupabaseClient into service, call async method, await Future.", "Mock response JSON", "Service parses mocked API response correctly into typed domain objects."),
        ("Verify error boundary & fallback widget rendering on exception", "Trigger exception during widget build phase inside ErrorWidget boundary.", "Thrown Exception()", "ErrorWidget fallback renders gracefully without throwing unhandled Flutter error.")
    ]

    for i in range(1, 61):
        tc_id = f"FLU-TST-{i:03d}"
        mod = flutter_modules[i-1]
        scen_idx = (i - 1) % len(flutter_scenarios)
        template = flutter_scenarios[scen_idx]

        title = f"Flutter Test - {mod} - {template[0]} (#{i})"
        scenario = f"Unit/Widget testing of {mod} components"
        obj = f"Verify isolated unit logic and widget tree integrity for {mod} using flutter_test package."
        pre = f"Flutter SDK configured, flutter_test runner initialized with mock dependencies."
        steps = f"1. Initialize mock dependencies.\n2. {template[1]}\n3. Verify test assertions and expectations."
        data = f"Target Component: {mod}, Framework: Flutter 3.x WidgetTester"
        exp = f"{template[3]} in unit test suite."
        prio = "High" if i % 4 == 0 else "Medium"
        sev = "Critical" if i % 10 == 0 else "Moderate"

        add_tc(tc_id, "Flutter Unit and Widget Testing", mod, scenario, title, obj, pre, steps, data, exp, prio, sev)

    # =========================================================================
    # 4. Functional and Integration Testing (50 Test Cases: FUN-INT-001 to FUN-INT-050)
    # =========================================================================
    functional_modules = (MODULES * 2)[:50]

    functional_scenarios = [
        ("Verify end-to-end data flow between client UI and Supabase DB", "Perform data creation in client UI, inspect network payload, verify DB record.", "User profile payload", "Record inserted cleanly into Supabase database with matching user_id."),
        ("Verify state synchronization across multi-screen workflow", "Update preference in Settings, navigate to Dashboard, check updated state.", "User preference key", "Dashboard reflects updated setting immediately without manual refresh."),
        ("Verify transactional workflow completion (Order / Workout log)", "Complete multi-step form workflow, submit transaction, verify summary page.", "Transaction payload", "Transaction completes successfully, DB state updated, confirmation displayed."),
        ("Verify graceful handling of network disconnect and reconnect sync", "Simulate offline state during data edit, restore network connection.", "Network drop simulation", "App queues unsynced changes locally and syncs with server upon connection."),
        ("Verify role-based access control and RLS query policies", "Attempt database access with unauthorized user token.", "Anon vs Authenticated token", "Supabase RLS rejects unauthorized query with HTTP 403 / permission error.")
    ]

    for i in range(1, 51):
        tc_id = f"FUN-INT-{i:03d}"
        mod = functional_modules[i-1]
        scen_idx = (i - 1) % len(functional_scenarios)
        template = functional_scenarios[scen_idx]

        title = f"Functional Integration - {mod} - {template[0]} (#{i})"
        scenario = f"Integration testing of {mod} integration point"
        obj = f"Verify cross-module functionality and Supabase integration integrity for {mod}."
        pre = f"Test environment active, backend API / Supabase instance reachable."
        steps = f"1. Prepare test user session.\n2. Execute {mod} integration scenario.\n3. {template[1]}\n4. Validate end-to-end result."
        data = f"Module: {mod}, API Endpoint: Supabase Rest / Realtime API"
        exp = f"{template[3]} for module {mod}."
        prio = "High" if i % 2 == 0 else "Medium"
        sev = "Major" if i % 5 == 0 else "Moderate"

        add_tc(tc_id, "Functional and Integration Testing", mod, scenario, title, obj, pre, steps, data, exp, prio, sev)

    # =========================================================================
    # 5. Load and Performance Testing (30 Test Cases: PER-LOD-001 to PER-LOD-030)
    # =========================================================================
    load_modules = MODULES[:30]

    load_scenarios = [
        ("Verify REST API endpoint throughput under 100 concurrent virtual users", "Locust / JMeter load test script executing against localhost test environment.", "100 VU, Hatch rate 10/s", "Response time p95 < 300ms, error rate 0.0%."),
        ("Verify UI rendering framerate (60 FPS) during rapid scrolling animation", "Scroll list view continuously for 60 seconds while monitoring Flutter Performance overlay.", "100 list items with images", "Frame rendering stays steady at 60 FPS without jank or dropped frames."),
        ("Verify memory consumption stability during prolonged application usage", "Perform continuous navigation transitions across views for 15 minutes.", "Memory Profiler session", "Heap memory remains stable under 180MB without progressive memory leaks."),
        ("Verify database query response latency under bulk data insertion", "Insert 500 records concurrently into Supabase table in test sandbox.", "500 mock records", "Total insertion batch time < 1.5 seconds."),
        ("Verify asset image loading optimization and caching performance", "Load media gallery containing 50 compressed webp assets.", "WebP asset suite", "First contentful paint < 400ms, cached reload < 50ms.")
    ]

    for i in range(1, 31):
        tc_id = f"PER-LOD-{i:03d}"
        mod = load_modules[i-1]
        scen_idx = (i - 1) % len(load_scenarios)
        template = load_scenarios[scen_idx]

        title = f"Performance - {mod} - {template[0]} (#{i})"
        scenario = f"Load & Performance evaluation of {mod}"
        obj = f"Ensure {mod} meets strict performance thresholds under load on local test server."
        pre = f"Local test server active on localhost:8080 / Supabase local Docker stack."
        steps = f"1. Spin up performance benchmark tools.\n2. Target {mod} endpoint/component.\n3. {template[1]}\n4. Collect latency/FPS telemetry."
        data = f"Target: http://localhost:8080/api/{mod.lower().replace(' ', '_')}, Load Target: {template[2]}"
        exp = f"{template[3]} for {mod}."
        prio = "High" if i % 3 == 0 else "Medium"
        sev = "Major" if i % 4 == 0 else "Moderate"

        add_tc(tc_id, "Load and Performance Testing", mod, scenario, title, obj, pre, steps, data, exp, prio, sev)

    # =========================================================================
    # 6. Safe Vulnerability and Configuration Checks (30 Test Cases: SEC-VUL-001 to SEC-VUL-030)
    # =========================================================================
    sec_modules = MODULES[:30]

    sec_scenarios = [
        ("Verify environment variable isolation and absence of hardcoded secrets", "Scan repository codebase and build artifacts for plain-text API keys, DB passwords, or credentials.", ".env vs source code scan", "Zero sensitive credentials found hardcoded in source files or client assets."),
        ("Verify Row Level Security (RLS) enforcement on Supabase tables", "Attempt SELECT/UPDATE query on private user records using secondary user JWT auth token.", "Supabase client query with Auth token B", "Query returns 0 records or permission error; access strictly isolated by user_id."),
        ("Verify HTTPS/TLS communication enforcement for all remote API calls", "Audit network requests originating from client application.", "OWASP ZAP passive proxy audit", "All outbound connections enforce TLS 1.3/1.2; insecure HTTP rejected."),
        ("Verify SQL Injection sanitization in search filter input fields", "Submit payload `' OR '1'='1` in search inputs across screens.", "Safe SQL injection payload", "Input sanitized properly; no SQL syntax error or unauthorized data disclosure."),
        ("Verify XSS (Cross-Site Scripting) prevention in user input fields", "Submit payload `<script>alert('xss')</script>` in user text input fields.", "Safe XSS HTML payload", "Input escaped safely as literal string; script tag not executed in DOM context.")
    ]

    for i in range(1, 31):
        tc_id = f"SEC-VUL-{i:03d}"
        mod = sec_modules[i-1]
        scen_idx = (i - 1) % len(sec_scenarios)
        template = sec_scenarios[scen_idx]

        title = f"Security Check - {mod} - {template[0]} (#{i})"
        scenario = f"Safe configuration & security audit for {mod}"
        obj = f"Verify OWASP compliance and safe configuration posture for {mod} without external scanning."
        pre = f"Static analysis tool ready, test environment sandbox active locally."
        steps = f"1. Audit {mod} implementation configuration.\n2. {template[1]}\n3. Validate zero security vulnerabilities."
        data = f"Audit Scope: {mod}, Method: Safe passive check / Static analysis"
        exp = f"{template[3]} for {mod}."
        prio = "High"
        sev = "Critical" if i % 2 == 0 else "Major"

        add_tc(tc_id, "Safe Vulnerability and Configuration Checks", mod, scenario, title, obj, pre, steps, data, exp, prio, sev)

    return test_cases


# ----------------------------------------------------
# Main Excel Workbook Builder
# ----------------------------------------------------
def build_excel_report():
    all_cases = generate_all_360_test_cases()
    print(f"Total Test Cases Generated: {len(all_cases)}")

    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Styles Setup
    font_family = "Segoe UI"
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="10A866", end_color="10A866", fill_type="solid") # NutriFit Green
    
    title_font = Font(name=font_family, size=16, bold=True, color="0A5C36")
    subtitle_font = Font(name=font_family, size=10, italic=True, color="555555")
    section_font = Font(name=font_family, size=12, bold=True, color="0A5C36")

    thin_border = Border(
        left=Side(style='thin', color='E0E0E0'),
        right=Side(style='thin', color='E0E0E0'),
        top=Side(style='thin', color='E0E0E0'),
        bottom=Side(style='thin', color='E0E0E0')
    )

    card_border = Border(
        left=Side(style='medium', color='10A866'),
        right=Side(style='medium', color='10A866'),
        top=Side(style='medium', color='10A866'),
        bottom=Side(style='medium', color='10A866')
    )

    status_styles = {
        "Passed": (PatternFill(start_color="E8F8F0", end_color="E8F8F0", fill_type="solid"), Font(name=font_family, color="10A866", bold=True)),
        "Failed": (PatternFill(start_color="FCE8E6", end_color="FCE8E6", fill_type="solid"), Font(name=font_family, color="D93025", bold=True)),
        "Skipped": (PatternFill(start_color="FEF7E0", end_color="FEF7E0", fill_type="solid"), Font(name=font_family, color="B06000", bold=True)),
        "Blocked": (PatternFill(start_color="FFEFE6", end_color="FFEFE6", fill_type="solid"), Font(name=font_family, color="D9530F", bold=True)),
        "Not Executed": (PatternFill(start_color="F1F3F4", end_color="F1F3F4", fill_type="solid"), Font(name=font_family, color="5F6368", bold=True))
    }

    # Helper function to style detail sheet
    def create_detail_sheet(title, cases):
        ws = wb.create_sheet(title=title)
        ws.views.sheetView[0].showGridLines = True

        # Header Row
        ws.append(COLUMNS)
        ws.row_dimensions[1].height = 28
        ws.freeze_panes = "A2"

        for col_idx in range(1, len(COLUMNS) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Data Rows
        for row_idx, tc in enumerate(cases, start=2):
            row_data = [tc[col] for col in COLUMNS]
            ws.append(row_data)
            ws.row_dimensions[row_idx].height = 36 # Room for detailed text

            for col_idx in range(1, len(COLUMNS) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = Font(name=font_family, size=9)
                cell.border = thin_border
                cell.alignment = Alignment(vertical='top', wrap_text=True)

                # Center short columns
                if col_idx in [1, 2, 12, 13, 14]:
                    cell.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)

                # Highlight Status column (14th column)
                if col_idx == 14:
                    status_val = cell.value
                    if status_val in status_styles:
                        fill, font = status_styles[status_val]
                        cell.fill = fill
                        cell.font = font

        # Enable Autofilter
        ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{len(cases) + 1}"

        # Column Auto-Widths with smart max width
        col_width_overrides = {
            "Test Case ID": 14,
            "Testing Type": 26,
            "Module": 20,
            "Test Scenario": 30,
            "Test Case Title": 35,
            "Objective": 35,
            "Preconditions": 30,
            "Test Steps": 45,
            "Test Data": 25,
            "Expected Result": 35,
            "Actual Result": 20,
            "Priority": 12,
            "Severity": 12,
            "Status": 15,
            "Evidence Path": 25,
            "Remarks": 25
        }

        for col_idx, col_name in enumerate(COLUMNS, start=1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = col_width_overrides.get(col_name, 20)

    # ----------------------------------------------------
    # Sheet 1: Test Summary Dashboard
    # ----------------------------------------------------
    ws_sum = wb.create_sheet(title="Test Summary")
    ws_sum.views.sheetView[0].showGridLines = True

    # Title Block
    ws_sum["A1"] = "NutriFit Application - Complete QA Test Case Report"
    ws_sum["A1"].font = title_font
    ws_sum["A2"] = "Automated & Manual Quality Assurance Suite | Target: Web & Android | Total Test Cases: 360"
    ws_sum["A2"].font = subtitle_font

    ws_sum.row_dimensions[1].height = 25
    ws_sum.row_dimensions[2].height = 18

    # KPI Summary Cards (Rows 4 to 6)
    kpi_headers = ["Total Test Cases", "Passed", "Failed", "Skipped", "Blocked", "Not Executed", "Execution Progress"]
    kpi_formulas = [
        "=SUM(B9:B14)",
        '=COUNTIF(\'Selenium Tests\'!N:N, "Passed") + COUNTIF(\'Appium E2E Tests\'!N:N, "Passed") + COUNTIF(\'Flutter Tests\'!N:N, "Passed") + COUNTIF(\'Functional Tests\'!N:N, "Passed") + COUNTIF(\'Load Tests\'!N:N, "Passed") + COUNTIF(\'Safe Vulnerability Checks\'!N:N, "Passed")',
        '=COUNTIF(\'Selenium Tests\'!N:N, "Failed") + COUNTIF(\'Appium E2E Tests\'!N:N, "Failed") + COUNTIF(\'Flutter Tests\'!N:N, "Failed") + COUNTIF(\'Functional Tests\'!N:N, "Failed") + COUNTIF(\'Load Tests\'!N:N, "Failed") + COUNTIF(\'Safe Vulnerability Checks\'!N:N, "Failed")',
        '=COUNTIF(\'Selenium Tests\'!N:N, "Skipped") + COUNTIF(\'Appium E2E Tests\'!N:N, "Skipped") + COUNTIF(\'Flutter Tests\'!N:N, "Skipped") + COUNTIF(\'Functional Tests\'!N:N, "Skipped") + COUNTIF(\'Load Tests\'!N:N, "Skipped") + COUNTIF(\'Safe Vulnerability Checks\'!N:N, "Skipped")',
        '=COUNTIF(\'Selenium Tests\'!N:N, "Blocked") + COUNTIF(\'Appium E2E Tests\'!N:N, "Blocked") + COUNTIF(\'Flutter Tests\'!N:N, "Blocked") + COUNTIF(\'Functional Tests\'!N:N, "Blocked") + COUNTIF(\'Load Tests\'!N:N, "Blocked") + COUNTIF(\'Safe Vulnerability Checks\'!N:N, "Blocked")',
        '=COUNTIF(\'Selenium Tests\'!N:N, "Not Executed") + COUNTIF(\'Appium E2E Tests\'!N:N, "Not Executed") + COUNTIF(\'Flutter Tests\'!N:N, "Not Executed") + COUNTIF(\'Functional Tests\'!N:N, "Not Executed") + COUNTIF(\'Load Tests\'!N:N, "Not Executed") + COUNTIF(\'Safe Vulnerability Checks\'!N:N, "Not Executed")',
        '=IF(B4=0, "0%", TEXT((B4-G4)/B4, "0.0%"))'
    ]

    for col_idx, (hdr, form) in enumerate(zip(kpi_headers, kpi_formulas), start=1):
        cell_hdr = ws_sum.cell(row=4, column=col_idx, value=hdr)
        cell_hdr.font = Font(name=font_family, size=9, bold=True, color="555555")
        cell_hdr.alignment = Alignment(horizontal='center', vertical='center')
        cell_hdr.fill = PatternFill(start_color="F5F7FA", end_color="F5F7FA", fill_type="solid")

        cell_val = ws_sum.cell(row=5, column=col_idx, value=form)
        cell_val.font = Font(name=font_family, size=14, bold=True, color="0A5C36")
        cell_val.alignment = Alignment(horizontal='center', vertical='center')
        cell_val.border = card_border

    # Section 1: Testing Type Summary Table (Row 8)
    ws_sum["A7"] = "1. Testing Type Breakdown"
    ws_sum["A7"].font = section_font

    tt_headers = ["Testing Type", "Total Cases", "Passed", "Failed", "Skipped", "Blocked", "Not Executed", "Pass Rate %"]
    ws_sum.append(tt_headers)
    ws_sum.row_dimensions[8].height = 24

    for col_idx in range(1, len(tt_headers) + 1):
        cell = ws_sum.cell(row=8, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    tt_sheet_map = [
        ("Selenium Web UI Testing", "Selenium Tests", 100),
        ("Appium Android E2E Testing", "Appium E2E Tests", 90),
        ("Flutter Unit and Widget Testing", "Flutter Tests", 60),
        ("Functional and Integration Testing", "Functional Tests", 50),
        ("Load and Performance Testing", "Load Tests", 30),
        ("Safe Vulnerability and Configuration Checks", "Safe Vulnerability Checks", 30),
    ]

    for r_idx, (tt_name, sh_name, count) in enumerate(tt_sheet_map, start=9):
        ws_sum.cell(row=r_idx, column=1, value=tt_name)
        ws_sum.cell(row=r_idx, column=2, value=count)
        ws_sum.cell(row=r_idx, column=3, value=f'=COUNTIF(\'{sh_name}\'!N:N, "Passed")')
        ws_sum.cell(row=r_idx, column=4, value=f'=COUNTIF(\'{sh_name}\'!N:N, "Failed")')
        ws_sum.cell(row=r_idx, column=5, value=f'=COUNTIF(\'{sh_name}\'!N:N, "Skipped")')
        ws_sum.cell(row=r_idx, column=6, value=f'=COUNTIF(\'{sh_name}\'!N:N, "Blocked")')
        ws_sum.cell(row=r_idx, column=7, value=f'=COUNTIF(\'{sh_name}\'!N:N, "Not Executed")')
        ws_sum.cell(row=r_idx, column=8, value=f'=IF(B{r_idx}=0, "0%", TEXT(C{r_idx}/B{r_idx}, "0.0%"))')

        for c_idx in range(1, 9):
            cell = ws_sum.cell(row=r_idx, column=c_idx)
            cell.font = Font(name=font_family, size=10)
            cell.border = thin_border
            if c_idx >= 2:
                cell.alignment = Alignment(horizontal='center', vertical='center')

    # Total Row for Testing Types
    ws_sum.cell(row=15, column=1, value="Total Suite").font = Font(name=font_family, size=10, bold=True)
    ws_sum.cell(row=15, column=2, value="=SUM(B9:B14)").font = Font(name=font_family, size=10, bold=True)
    ws_sum.cell(row=15, column=3, value="=SUM(C9:C14)").font = Font(name=font_family, size=10, bold=True)
    ws_sum.cell(row=15, column=4, value="=SUM(D9:D14)").font = Font(name=font_family, size=10, bold=True)
    ws_sum.cell(row=15, column=5, value="=SUM(E9:E14)").font = Font(name=font_family, size=10, bold=True)
    ws_sum.cell(row=15, column=6, value="=SUM(F9:F14)").font = Font(name=font_family, size=10, bold=True)
    ws_sum.cell(row=15, column=7, value="=SUM(G9:G14)").font = Font(name=font_family, size=15, bold=True)
    ws_sum.cell(row=15, column=8, value='=IF(B15=0, "0%", TEXT(C15/B15, "0.0%"))').font = Font(name=font_family, size=10, bold=True)

    for c_idx in range(1, 9):
        cell = ws_sum.cell(row=15, column=c_idx)
        cell.border = thin_border
        cell.fill = PatternFill(start_color="F0F9F4", end_color="F0F9F4", fill_type="solid")
        if c_idx >= 2:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # Section 2: Module-Wise Summary Table (Row 17)
    ws_sum["A17"] = "2. Module-Wise Test Coverage Summary (All 37 Modules)"
    ws_sum["A17"].font = section_font

    mod_headers = ["Module Name", "Total Cases", "Passed", "Failed", "Skipped", "Blocked", "Not Executed"]
    ws_sum.cell(row=18, column=1, value=mod_headers[0])
    for idx, h in enumerate(mod_headers[1:], start=2):
        ws_sum.cell(row=18, column=idx, value=h)

    for c_idx in range(1, 8):
        cell = ws_sum.cell(row=18, column=c_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for r_offset, mod_name in enumerate(MODULES, start=19):
        ws_sum.cell(row=r_offset, column=1, value=mod_name)
        
        # SUM COUNTIFS across all 6 test detail sheets
        tot_formula = "+".join([f'COUNTIF(\'{sh}\'!C:C, A{r_offset})' for _, sh, _ in tt_sheet_map])
        pass_formula = "+".join([f'COUNTIFS(\'{sh}\'!C:C, A{r_offset}, \'{sh}\'!N:N, "Passed")' for _, sh, _ in tt_sheet_map])
        fail_formula = "+".join([f'COUNTIFS(\'{sh}\'!C:C, A{r_offset}, \'{sh}\'!N:N, "Failed")' for _, sh, _ in tt_sheet_map])
        skip_formula = "+".join([f'COUNTIFS(\'{sh}\'!C:C, A{r_offset}, \'{sh}\'!N:N, "Skipped")' for _, sh, _ in tt_sheet_map])
        block_formula = "+".join([f'COUNTIFS(\'{sh}\'!C:C, A{r_offset}, \'{sh}\'!N:N, "Blocked")' for _, sh, _ in tt_sheet_map])
        not_exec_formula = "+".join([f'COUNTIFS(\'{sh}\'!C:C, A{r_offset}, \'{sh}\'!N:N, "Not Executed")' for _, sh, _ in tt_sheet_map])

        ws_sum.cell(row=r_offset, column=2, value=f"={tot_formula}")
        ws_sum.cell(row=r_offset, column=3, value=f"={pass_formula}")
        ws_sum.cell(row=r_offset, column=4, value=f"={fail_formula}")
        ws_sum.cell(row=r_offset, column=5, value=f"={skip_formula}")
        ws_sum.cell(row=r_offset, column=6, value=f"={block_formula}")
        ws_sum.cell(row=r_offset, column=7, value=f"={not_exec_formula}")

        for c_idx in range(1, 8):
            cell = ws_sum.cell(row=r_offset, column=c_idx)
            cell.font = Font(name=font_family, size=9)
            cell.border = thin_border
            if c_idx >= 2:
                cell.alignment = Alignment(horizontal='center', vertical='center')

    # Total Row for Modules
    tot_mod_row = 19 + len(MODULES)
    ws_sum.cell(row=tot_mod_row, column=1, value="Total All Modules").font = Font(name=font_family, size=10, bold=True)
    for c_idx in range(2, 8):
        col_let = get_column_letter(c_idx)
        cell = ws_sum.cell(row=tot_mod_row, column=c_idx, value=f"=SUM({col_let}19:{col_let}{tot_mod_row-1})")
        cell.font = Font(name=font_family, size=10, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        cell.fill = PatternFill(start_color="F0F9F4", end_color="F0F9F4", fill_type="solid")

    ws_sum.column_dimensions['A'].width = 38
    ws_sum.column_dimensions['B'].width = 16
    ws_sum.column_dimensions['C'].width = 14
    ws_sum.column_dimensions['D'].width = 14
    ws_sum.column_dimensions['E'].width = 14
    ws_sum.column_dimensions['F'].width = 14
    ws_sum.column_dimensions['G'].width = 16
    ws_sum.column_dimensions['H'].width = 16

    # ----------------------------------------------------
    # Detail Sheets Generation (2 to 7)
    # ----------------------------------------------------
    sel_cases = [tc for tc in all_cases if tc["Testing Type"] == "Selenium Web UI Testing"]
    app_cases = [tc for tc in all_cases if tc["Testing Type"] == "Appium Android E2E Testing"]
    flu_cases = [tc for tc in all_cases if tc["Testing Type"] == "Flutter Unit and Widget Testing"]
    fun_cases = [tc for tc in all_cases if tc["Testing Type"] == "Functional and Integration Testing"]
    lod_cases = [tc for tc in all_cases if tc["Testing Type"] == "Load and Performance Testing"]
    sec_cases = [tc for tc in all_cases if tc["Testing Type"] == "Safe Vulnerability and Configuration Checks"]

    create_detail_sheet("Selenium Tests", sel_cases)
    create_detail_sheet("Appium E2E Tests", app_cases)
    create_detail_sheet("Flutter Tests", flu_cases)
    create_detail_sheet("Functional Tests", fun_cases)
    create_detail_sheet("Load Tests", lod_cases)
    create_detail_sheet("Safe Vulnerability Checks", sec_cases)

    # ----------------------------------------------------
    # Sheet 8: Failed Tests Sheet
    # ----------------------------------------------------
    ws_fail = wb.create_sheet(title="Failed Tests")
    ws_fail.views.sheetView[0].showGridLines = True
    ws_fail.append(COLUMNS)
    ws_fail.row_dimensions[1].height = 28
    ws_fail.freeze_panes = "A2"

    for col_idx in range(1, len(COLUMNS) + 1):
        cell = ws_fail.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = PatternFill(start_color="D93025", end_color="D93025", fill_type="solid") # Red Header for Failed sheet
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Add explanatory baseline row if no failures initially
    ws_fail.append([
        "N/A", "System Filter", "All Modules", "Failed Test Tracking View",
        "No Failed Test Cases Recorded", "Track test execution failures dynamically",
        "Initial status baseline: All test cases set to 'Not Executed'",
        "1. Execute automated CI test suite.\n2. Filter test cases with Status == 'Failed'.\n3. Populate failure log here.",
        "Status: Not Executed", "0 Failures detected in initial baseline run",
        "All test cases currently Not Executed", "High", "Critical", "Not Executed",
        "N/A", "This sheet serves as the dedicated failure review log during test execution cycles."
    ])

    for col_idx in range(1, len(COLUMNS) + 1):
        cell = ws_fail.cell(row=2, column=col_idx)
        cell.font = Font(name=font_family, size=9)
        cell.border = thin_border
        cell.alignment = Alignment(vertical='top', wrap_text=True)

    ws_fail.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}2"
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        ws_fail.column_dimensions[get_column_letter(col_idx)].width = 25

    # ----------------------------------------------------
    # Sheet 9: Execution Evidence Matrix
    # ----------------------------------------------------
    ws_evid = wb.create_sheet(title="Execution Evidence")
    ws_evid.views.sheetView[0].showGridLines = True
    evid_columns = ["Test Case ID", "Testing Type", "Module", "Test Scenario", "Status", "Evidence File Path", "Timestamp", "Artifact Name", "Verification Notes"]
    ws_evid.append(evid_columns)
    ws_evid.row_dimensions[1].height = 28
    ws_evid.freeze_panes = "A2"

    for col_idx in range(1, len(evid_columns) + 1):
        cell = ws_evid.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Add sample evidence records mapping to test cases
    for idx, tc in enumerate(all_cases[:20], start=2): # Baseline 20 evidence slots
        row_ev = [
            tc["Test Case ID"],
            tc["Testing Type"],
            tc["Module"],
            tc["Test Scenario"],
            tc["Status"],
            f"testing-reports/evidence/{tc['Test Case ID'].lower()}_log.png",
            "2026-07-21 23:30:00 UTC",
            f"nutrifit-{tc['Test Case ID'].lower()}-artifact.zip",
            "Baseline screenshot & DOM trace artifact allocated"
        ]
        ws_evid.append(row_ev)

        for c_idx in range(1, len(evid_columns) + 1):
            cell = ws_evid.cell(row=idx, column=c_idx)
            cell.font = Font(name=font_family, size=9)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='top', wrap_text=True)

    ws_evid.auto_filter.ref = f"A1:{get_column_letter(len(evid_columns))}{21}"
    for col_idx in range(1, len(evid_columns) + 1):
        ws_evid.column_dimensions[get_column_letter(col_idx)].width = 24

    # Save Workbook
    out_path = "testing-reports/NutriFit_360_Test_Cases.xlsx"
    wb.save(out_path)
    print(f"SUCCESS: Excel workbook saved at '{out_path}'")

if __name__ == "__main__":
    build_excel_report()
